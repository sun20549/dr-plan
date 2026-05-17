#!/usr/bin/env python3
"""
TWLife 自動驗證腳本 — 從 .xlsm「明細版」cached values 直接對照網頁計算結果

用法:
  python3 tools/validate_twlife.py

每個 .xlsm 用「Excel 最後一次儲存時的快取參數」(性別/年齡/面額/年期)做驗證:
  1. 讀「要保資料輸入」抓 input
  2. 讀「明細版」抓每年的試算結果
  3. 用 Python port 跑等效計算
  4. 印 Δ 對照表

要驗其他參數 → 你在 Excel 開檔 → 改參數 → 存 → 重跑此腳本
"""
import msoffcrypto, io, json, sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path('/sessions/gallant-kind-faraday/mnt/台灣')  # 你的 Excel 資料夾(mounted)
DATA = Path(__file__).resolve().parent.parent / 'data' / 'twlife'

# (xlsm 檔名前綴 → plan_code)
PRODUCTS = {
    "台灣人壽臻威豐": "TLZWF6",
    "台灣人壽美世長紅": "TLMSCH06",
    "台灣人壽美紅旺": "TLMHW06",
    # 待驗證(catalog 暫不上):
    # "台灣人壽美紅勝利": "TLMHSL01/02",
    # "台灣人壽美紅富利": "TLMHFL01",
    # "台灣人壽美紅鑽": "TLMHZ02",
}


def open_xlsm(path, pwd='0800099850'):
    raw = open(path, 'rb').read()
    try:
        o = msoffcrypto.OfficeFile(io.BytesIO(raw))
        if o.is_encrypted():
            o.load_key(password=pwd)
            out = io.BytesIO()
            o.decrypt(out)
            out.seek(0)
            return load_workbook(out, read_only=True, data_only=True, keep_vba=False)
    except Exception:
        pass
    return load_workbook(io.BytesIO(raw), read_only=True, data_only=True, keep_vba=False)


def read_xlsm_cached(path):
    """讀 .xlsm 的 input + 明細版每年 row"""
    wb = open_xlsm(path)
    ws_in = wb["要保資料輸入"]
    inputs = {c.coordinate: c.value for r in ws_in.iter_rows(min_row=1, max_row=44)
              for c in r if c.value is not None}

    sex_zh = inputs.get('D9', '')
    sex = 'M' if '男' in str(sex_zh) else 'F'
    age = int(inputs.get('D15') or inputs.get('D19') or 0)
    face = int(inputs.get('D22') or 0)
    pay_yr = int(inputs.get('D23') or 6)

    # 明細版 row 19 起,col B=yr, C=attained, D=cumP, W=身故, T=解約, AA=翌日解約
    rows = []
    if "明細版" in wb.sheetnames:
        ws = wb["明細版"]
        last_yr = 0
        for r in range(19, ws.max_row + 1):
            row = ws[r]
            yr = row[1].value
            if not isinstance(yr, (int, float)):
                continue
            yr = int(yr)
            if yr < last_yr - 5:  # 換情境了
                break
            last_yr = yr
            rows.append({
                'yr': yr,
                'attained': row[2].value,
                'cumPrem': row[3].value,
                'J': row[22].value,    # W 身故
                'K': row[26].value,    # AA 翌日解約 (含 TDS)
            })
    return {'sex': sex, 'age': age, 'face': face, 'pay_yr': pay_yr}, rows


def calc_twlife(plan_code, sex, age, face, pay_yr):
    """Python port of calculateTWLife()"""
    json_path = next(DATA.glob(f"{plan_code}_*.json"))
    data = json.load(open(json_path))
    premrate = data['gp'][plan_code][sex].get(str(age))
    if premrate is None:
        return []
    ageU = data['uv'][plan_code][sex][str(age)]
    ageD = data['div'][plan_code][sex][str(age)]
    crit = data['corridor_criteria']
    tws = data['twlife_specific']
    apv = tws['addPolicyValues'][sex][str(age)]
    acv = tws['addCashValues'][sex][str(age)]
    pws = tws['pws']
    pb = tws.get('premium_base', 1.06)

    raw = round(face * premrate / 1000)
    if raw < 10000: dh = 0
    elif raw < 14000: dh = 0.01
    elif raw < 18000: dh = 0.02
    else: dh = 0.03
    first_per = round(raw * (1 - dh - 0.01))
    renew_per = round(raw * (1 - dh - 0.01))

    mat_yr = 111 - age
    nfv110 = ageU[mat_yr][0] / 1000 if mat_yr < len(ageU) else 0
    cumP = cumE = cumRaw = 0
    rows = []
    last_yr = 110 - age + 1
    for yr in range(1, last_yr + 1):
        is_pay = yr <= pay_yr
        A = (first_per if yr == 1 else renew_per) if is_pay else 0
        cumP += A
        if is_pay: cumRaw += raw
        atn = age + yr - 1
        cv = crit.get(str(min(atn, 110)), 1)
        has_pws = bool(pws)
        pws_yr = (pws.get(str(yr), 1.0) if has_pws else 0) if yr >= pay_yr else 0
        NFV = ageU[yr][0] if yr < len(ageU) else 0
        CSV = ageU[yr][1] if yr < len(ageU) else 0
        apvY = apv[yr] if yr < len(apv) else 0
        acvY = acv[yr] if yr < len(acv) else 0

        C_p = cumRaw * pb
        C_pv = NFV * cv * face / 1000
        C_pws = pws_yr * face
        C_n110 = nfv110 * face if yr >= pay_yr else 0
        C, w = C_p, 'p'
        if C_pv > C: C, w = C_pv, 'n'
        if C_pws > C: C, w = C_pws, 'w'
        if C_n110 > C: C, w = C_n110, '110'

        D = CSV * face / 1000
        div = ageD[yr] if yr < len(ageD) else [0, 0, 0]
        AD = div[0] * face / 10000
        bought = AD / apvY if apvY > 0 else 0
        cumE += round(bought)
        if w == 'p': F = cumE * C_p / face
        elif w == 'n': F = cumE * apvY * cv
        elif w == 'w': F = cumE * pws_yr
        else: F = cumE * nfv110
        G = cumE * acvY
        TDD = div[1] * face
        TDS = div[2] * face
        J = C + F + TDD
        K = D + G + TDS
        if atn >= 110: K = J
        rows.append({'yr': yr, 'attained': atn, 'cumP': cumP, 'J': J, 'K': K})
    return rows


def validate(prefix, plan_code):
    files = list(ROOT.glob(f"{prefix}*.xlsm"))
    if not files:
        print(f"[{plan_code}] ✗ 找不到檔案 {prefix}*.xlsm")
        return
    src = sorted(files)[0]
    print(f"\n{'='*100}\n[{plan_code}] {src.name}\n{'='*100}")
    try:
        params, excel = read_xlsm_cached(src)
    except Exception as e:
        print(f"  ✗ 讀 Excel 失敗:{e}")
        return
    print(f"  Excel cached input: {params['sex']}{params['age']}歲 face {params['face']:,} USD {params['pay_yr']}年期")
    print(f"  Excel 明細版列數:{len(excel)}")
    if not excel:
        print(f"  ⚠ 明細版無資料(可能 Excel 還沒按過試算)")
        return

    calc = calc_twlife(plan_code, params['sex'], params['age'], params['face'], params['pay_yr'])
    print(f"  Calc 列數:{len(calc)}")
    if not calc:
        print(f"  ✗ 計算失敗(可能此 sex/age 沒費率)")
        return

    sample_yrs = [1, 2, 3, 6, 10, 15, 20, 30, 50, len(excel)-1] if len(excel) > 50 else [1, 2, 3, 6, 10, len(excel)-1]
    print(f"\n  {'yr':>3} {'atn':>3} | {'Excel cumP':>12} {'Calc cumP':>12} {'ΔP':>6} | {'Excel J':>13} {'Calc J':>13} {'ΔJ':>9} | {'Excel K':>13} {'Calc K':>13} {'ΔK':>9}")
    print('  ' + '-' * 130)
    ok_J = ok_K = 0; total = 0
    for yr in sample_yrs:
        if yr < 1 or yr > min(len(excel), len(calc)): continue
        e, c = excel[yr-1], calc[yr-1]
        ev_J = e['J'] or 0; ev_K = e['K'] or 0; ev_P = e['cumPrem'] or 0
        dJ = c['J'] - ev_J; dK = c['K'] - ev_K; dP = c['cumP'] - ev_P
        if abs(dJ) < 2: ok_J += 1
        if abs(dK) < 2: ok_K += 1
        total += 1
        print(f"  {yr:>3} {e['attained']:>3} | {ev_P:>12,.0f} {c['cumP']:>12,.0f} {dP:>+6.0f} | {ev_J:>13,.2f} {c['J']:>13,.2f} {dJ:>+9.2f} | {ev_K:>13,.2f} {c['K']:>13,.2f} {dK:>+9.2f}")
    print(f"\n  匹配率 (Δ<2 USD):  J身故 {ok_J}/{total}   K解約 {ok_K}/{total}")


if __name__ == "__main__":
    for prefix, code in PRODUCTS.items():
        validate(prefix, code)
