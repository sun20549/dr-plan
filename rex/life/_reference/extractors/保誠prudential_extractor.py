#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
prudential_extractor.py — 保誠 Excel → JSON 抽取器(v3.1 規則完整實作)

支援:
  - prudential_v2 引擎(layout A 85 欄 / layout B 67 欄分區結構)
  - .xls 自動透過 LibreOffice 轉檔
  - 檔名 vs Excel 內容驗證(防止業務員整理檔案時複製錯內容)
  - 重複版本檢測(較新覆蓋較舊,較舊不蓋較新)
  - 13 條自洽性驗證(含身故倍率寬鬆規則、分紅生效時間寬鬆規則)
  - 批次模式(單檔失敗繼續處理下一個)

用法:
  python prudential_extractor.py 檔案.xls
  python prudential_extractor.py 檔案.xls --output ./out/
  python prudential_extractor.py *.xls --batch
  python prudential_extractor.py 檔案.xls --check-only         # 只驗檔名 vs 內容,不抽
  python prudential_extractor.py 檔案.xls --force-overwrite    # 跳過版本檢查
  python prudential_extractor.py 檔案.xls --skip-validation    # 跳過自洽性驗證(慎用)
  python prudential_extractor.py 檔案.xls -v                   # 詳細日誌

退出碼:
  0 = 成功
  1 = 驗證失敗(自洽性檢查 ❌)
  2 = 跳過(檔名錯 / 已有更新版本)
  3 = 不支援(非 prudential_v2 結構)
  4 = 系統錯誤(LibreOffice 失敗、檔不存在等)

依賴:
  pip install openpyxl
  apt install libreoffice  # .xls 轉檔用,沒裝就只能處理 .xlsx
"""
import argparse
import glob
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass, field, asdict
from datetime import date
from pathlib import Path
from typing import Any, Optional

import openpyxl


# ============================================================================
# 常數定義
# ============================================================================

VERSION = '3.1.0'

ENGINE = 'prudential_v2'
COMPANY = '保誠人壽'

# 資料輸入頁名稱變體(按優先順序找)
INPUT_SHEET_CANDIDATES = ['資料輸入', '資料輸入_主被保人']

# Layout A: 85 欄直線結構(ARLPLU30/57/64 / ARLPLN10/11/20 / ARLPLU63 等)
#   R4 col 2 = "保單年度",資料 R7 起,col 13/17/21 三情境直線排列
LAYOUT_A_COLS = {
    'header_row': 4,
    'data_start': 7,
    'y': 2, 'age': 3,
    'cum_prem': 5,         # 折扣後累積保費
    'cum_prem_gross': 6,    # 折扣前累積保費
    'death_benefit': 8,
    'cv_total': 9,
    'survival': 10,         # 通常 0(layout A 多為純壽險)
    'policy_value': 11,
    'reduced_paid_up': 12,
    'none_div_year': 13, 'none_div_cum_db': 15, 'none_div_cum_cv': 16,
    'mid_div_year':  17, 'mid_div_cum_db':  19, 'mid_div_cum_cv':  20,
    'low_div_year':  21, 'low_div_cum_db':  23, 'low_div_cum_cv':  24,
}

# Layout B: 67 欄「現金給付/儲存生息」分區結構(ACLPEU/ACLPEN 系列)
#   R2 col 1 = "年期",資料 R5 起
#   col 27/28/29 是「解約金 + 長青解約紅利」(中情境),非 col 33-38(那是含累積生存金)
#   col 30/31/32 是「壽險保障 + 長青紅利」
LAYOUT_B_COLS = {
    'header_row': 2,
    'data_start': 5,
    'y': 1, 'age': 2,
    'cum_prem': 5,          # 累積保費(已折扣)
    'survival': 6,          # 當年生存保險金(還本)
    'survival_cum': 7,
    'cv_total': 8,          # 解約金(無紅利 baseline)
    'death_benefit': 9,     # 壽險保障(無紅利 baseline)
    'policy_value': 22,
    'reduced_paid_up': 23,
    'none_div_year': 10, 'none_div_cum_cv': 27, 'none_div_cum_db': 30,
    'mid_div_year':  11, 'mid_div_cum_cv':  28, 'mid_div_cum_db':  31,
    'low_div_year':  12, 'low_div_cum_cv':  29, 'low_div_cum_db':  32,
}

# 預設輸出目錄
DEFAULT_OUTPUT_DIR = '/mnt/user-data/outputs'


# ============================================================================
# 工具函數
# ============================================================================

def f_num(v) -> Optional[float]:
    """正規化為數字。Excel 偶有 '1,234' 字串型數字。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(',', '')
    if s in ('', '-', 'N/A', 'na'):
        return None
    try:
        return float(s) if '.' in s else int(s)
    except ValueError:
        return None


def i_num(v) -> Optional[int]:
    n = f_num(v)
    return int(n) if n is not None else None


def find_label(ws, keyword: str, max_row: int = 100, max_col: int = 35):
    """搜尋含 keyword 的 cell,回傳 [(row, col, value)] 列表。"""
    hits = []
    for r in range(1, min(max_row, ws.max_row) + 1):
        for c in range(1, min(max_col, ws.max_column) + 1):
            v = ws.cell(r, c).value
            if v is not None and keyword in str(v):
                hits.append((r, c, v))
    return hits


def extract_date_from_filename(s: str) -> str:
    """從檔名/source_file 抽日期串(8 位數字),取最大那個。"""
    matches = re.findall(r'(20\d{6})', s)
    return max(matches) if matches else '00000000'


def extract_plan_code(fname: str) -> Optional[str]:
    """從檔名抽 plan_code(寬鬆 regex,接受 ARLPLU30 / ACLPEN26 / ARLPN003 等)。"""
    m = re.search(r'([A-Z]+\d+)', fname)
    return m.group(1) if m else None


def detect_currency(product_name: str) -> str:
    """從商品名推測幣別。Excel「幣別」欄不可靠。"""
    name = product_name or ''
    if any(kw in name for kw in ['外幣', '美元', 'USD', '美金']):
        return 'USD'
    if 'AUD' in name or '澳幣' in name:
        return 'AUD'
    return 'TWD'


# ============================================================================
# .xls → .xlsx 轉檔
# ============================================================================

def convert_xls_to_xlsx(xls_path: Path, verbose: bool = False) -> Path:
    """用 LibreOffice 轉 .xls 為 .xlsx,回傳 .xlsx 路徑。"""
    if xls_path.suffix.lower() == '.xlsx':
        return xls_path

    # 殺掉殘留 LibreOffice process(避免 lock)
    subprocess.run(['pkill', '-9', '-f', 'soffice'], capture_output=True)
    import time
    time.sleep(1)

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_xls = Path(tmpdir) / 'src.xls'
        shutil.copy(xls_path, tmp_xls)

        # 啟動 soffice,用 wait 等完成(不用 timeout 命令會被誤殺)
        result = subprocess.run(
            ['soffice', '--headless', '--convert-to', 'xlsx',
             '--outdir', tmpdir, str(tmp_xls)],
            capture_output=True, text=True, timeout=180
        )
        if verbose:
            print(f"  [LibreOffice] {result.stdout.strip()}", file=sys.stderr)

        out_xlsx = Path(tmpdir) / 'src.xlsx'
        if not out_xlsx.exists():
            raise RuntimeError(
                f"LibreOffice 轉檔失敗: {xls_path.name}\n"
                f"  stderr: {result.stderr[:300]}"
            )

        # 移到永久位置(/tmp 之外)
        final_path = xls_path.with_suffix('.xlsx')
        # 如果原始檔在 read-only 目錄,放到 /tmp
        try:
            shutil.copy(out_xlsx, final_path)
        except (PermissionError, OSError):
            final_path = Path(tempfile.mkdtemp()) / xls_path.with_suffix('.xlsx').name
            shutil.copy(out_xlsx, final_path)

        return final_path


# ============================================================================
# 結構與資料類別
# ============================================================================

@dataclass
class FileNameCheckResult:
    expected_code: Optional[str]
    actual_code: Optional[str]
    actual_product_name: Optional[str]
    matched: bool
    has_expected: bool
    has_actual: bool

    def report(self) -> str:
        lines = []
        lines.append(f"檔名代號: {self.expected_code or '(無)'}")
        lines.append(f"Excel R30 代號: {self.actual_code or '(找不到)'}")
        lines.append(f"Excel R1 商品名: {self.actual_product_name or '(找不到)'}")
        if self.matched:
            lines.append("結果: ✅ 對得上")
        elif not self.has_actual:
            lines.append("結果: ⚠️ 找不到 Excel 內代號(可能不是標準保誠檔)")
        else:
            lines.append(f"結果: ❌ 不符 → 內容是 {self.actual_code},不是 {self.expected_code}")
        return '\n'.join(lines)


@dataclass
class VersionCheckResult:
    plan_code: str
    output_dir: Path
    has_existing: bool
    existing_date: str
    new_date: str
    should_overwrite: bool

    def report(self) -> str:
        if not self.has_existing:
            return f"目錄沒有既有 {self.plan_code}.json,可寫入"
        prefix = "⚠️ 將覆蓋" if self.should_overwrite else "❌ 跳過"
        return (f"{prefix} {self.plan_code}.json "
                f"(既有 {self.existing_date} → 新 {self.new_date})")


@dataclass
class BaseParams:
    sex: Optional[str] = None
    age: Optional[int] = None
    sa: Optional[float] = None
    period: Optional[int] = None
    gross_premium: Optional[float] = None
    disc_premium: Optional[float] = None
    discount: Optional[float] = None
    discount_label: Optional[float] = None
    payout_period: Optional[int] = None
    product_name: Optional[str] = None

    def required_filled(self) -> list[str]:
        """v3.1 規則 7:必抽 4 個基準。"""
        missing = []
        if self.sex is None: missing.append('sex')
        if self.age is None: missing.append('age')
        if self.sa is None: missing.append('sa')
        if self.disc_premium is None: missing.append('disc_premium')
        return missing


# ============================================================================
# Step 0:結構偵察
# ============================================================================

def detect_layout(wb) -> Optional[str]:
    """回傳 'A' / 'B' / None"""
    if '試算表' not in wb.sheetnames:
        return None
    ws = wb['試算表']
    r4c2 = ws.cell(4, 2).value
    r2c1 = ws.cell(2, 1).value
    if r4c2 and '保單' in str(r4c2):
        return 'A'
    if r2c1 == '年期':
        return 'B'
    return None


def find_input_sheet(wb):
    """按 INPUT_SHEET_CANDIDATES 順序找資料輸入頁。"""
    for name in INPUT_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            return wb[name], name
    return None, None


def check_filename_vs_content(xlsx_path: Path, wb) -> FileNameCheckResult:
    """v3.1 Step 0.0:檔名 vs Excel 內容驗證。"""
    fname = xlsx_path.name
    expected = extract_plan_code(fname)

    ws_in, _ = find_input_sheet(wb)
    actual_code = None
    actual_name = None

    if ws_in:
        # R1.2 商品標題
        v = ws_in.cell(1, 2).value
        if v:
            actual_name = str(v).replace('_建議書', '').strip()

        # R28-40 找「主約商品」cell,該列 col 7 是代號
        for r in range(28, 40):
            for c in range(1, 6):
                if ws_in.cell(r, c).value == '主約商品':
                    code_v = ws_in.cell(r, 7).value
                    if code_v:
                        actual_code = str(code_v)
                    break
            if actual_code:
                break

    return FileNameCheckResult(
        expected_code=expected,
        actual_code=actual_code,
        actual_product_name=actual_name,
        matched=(expected is not None and actual_code == expected),
        has_expected=expected is not None,
        has_actual=actual_code is not None,
    )


def check_version_overwrite(plan_code: str, source_filename: str,
                            output_dir: Path) -> VersionCheckResult:
    """v3.1 Step 0.5:重複版本檢測。"""
    existing_path = output_dir / f'{plan_code}.json'
    if not existing_path.exists():
        return VersionCheckResult(
            plan_code=plan_code, output_dir=output_dir,
            has_existing=False, existing_date='', new_date='',
            should_overwrite=True,
        )

    with open(existing_path, encoding='utf-8') as f:
        existing = json.load(f)
    existing_src = existing.get('meta', {}).get('source_file', '')

    new_date = extract_date_from_filename(source_filename)
    old_date = extract_date_from_filename(existing_src)

    return VersionCheckResult(
        plan_code=plan_code, output_dir=output_dir,
        has_existing=True,
        existing_date=old_date, new_date=new_date,
        should_overwrite=(new_date > old_date),
    )


# ============================================================================
# Step 1:基準參數抽取
# ============================================================================

def extract_base_params(wb) -> BaseParams:
    """從資料輸入頁抽 base 參數。"""
    ws, sheet_name = find_input_sheet(wb)
    if ws is None:
        raise ValueError(
            f"找不到資料輸入頁。已知變體: {INPUT_SHEET_CANDIDATES}。"
            f"實際 sheets: {wb.sheetnames[:10]}"
        )

    base = BaseParams()

    # 1. 性別
    for r, c, _ in find_label(ws, '被保人'):
        for cc in range(c + 1, min(ws.max_column + 1, c + 14)):
            v = ws.cell(r, cc).value
            if v in ('男', '女'):
                base.sex = 'M' if v == '男' else 'F'
                break
        if base.sex:
            break

    # 2. 保險年齡
    for r, c, _ in find_label(ws, '保險年齡'):
        for cc in range(c + 1, min(ws.max_column + 1, c + 8)):
            v = ws.cell(r, cc).value
            if isinstance(v, (int, float)) and 0 < v < 120:
                base.age = int(v)
                break
        if base.age is not None:
            break

    # 3. 保額(嘗試多個 keyword)
    for kw in ['計算方式', '保險金額', '投保金額']:
        for r, c, _ in find_label(ws, kw):
            for cc in range(c + 1, min(ws.max_column + 1, c + 14)):
                v = ws.cell(r, cc).value
                if isinstance(v, (int, float)) and v >= 1000:
                    base.sa = v
                    break
            if base.sa is not None:
                break
        if base.sa is not None:
            break

    # 4. 繳費年期(從「主約商品」那列右邊找)
    for r, c, _ in find_label(ws, '主約商品'):
        for cc in range(c + 1, ws.max_column + 1):
            v = ws.cell(r, cc).value
            if isinstance(v, (int, float)) and 1 <= v <= 30:
                base.period = int(v)
                break
        if base.period is not None:
            break

    # 5. 折扣前/後總保費
    for r in range(1, min(ws.max_row + 1, 100)):
        for c in range(1, min(ws.max_column + 1, 30)):
            v = ws.cell(r, c).value
            if v is None:
                continue
            sv = str(v)
            if '原始總保費' in sv:
                for cc in range(c + 1, ws.max_column + 1):
                    vv = ws.cell(r, cc).value
                    if isinstance(vv, (int, float)) and vv > 0:
                        base.gross_premium = vv
                        break
            elif '折扣後總保費' in sv:
                for cc in range(c + 1, ws.max_column + 1):
                    vv = ws.cell(r, cc).value
                    if isinstance(vv, (int, float)) and vv > 0:
                        base.disc_premium = vv
                        break

    # 6. 折扣 = 1 - disc/gross
    if base.gross_premium and base.disc_premium and base.gross_premium > 0:
        base.discount = round(1 - base.disc_premium / base.gross_premium, 6)
    else:
        base.discount = 0

    # 7. 高保費折扣率(label,Excel 上明列的單一段折扣)
    for r, c, _ in find_label(ws, '高保費折扣率'):
        for rr in range(r, min(r + 5, ws.max_row + 1)):
            v = ws.cell(rr, c).value
            if isinstance(v, str):
                m = re.search(r'([\d.]+)\s*%', v)
                if m:
                    base.discount_label = float(m.group(1)) / 100
                    break
            elif isinstance(v, (int, float)) and 0 < v < 0.1:
                base.discount_label = v
                break
        if base.discount_label is not None:
            break

    # 8. payout_period(分期定期保險金給付期間)
    for kw in ['分期定期保險金給付期間', '給付期間']:
        for r, c, _ in find_label(ws, kw):
            for cc in range(c + 1, ws.max_column + 1):
                v = ws.cell(r, cc).value
                if isinstance(v, (int, float)) and 0 < v < 50:
                    base.payout_period = int(v)
                    break
            if base.payout_period is not None:
                break
        if base.payout_period is not None:
            break

    # 9. 商品名
    for r, c, _ in find_label(ws, '主約商品'):
        for cc in range(c + 1, ws.max_column + 1):
            v = ws.cell(r, cc).value
            if isinstance(v, str) and '保誠' in v and len(v) > 6:
                base.product_name = v
                break
        if base.product_name:
            break

    return base


# ============================================================================
# Step 2:Schedule 抽取
# ============================================================================

def extract_schedule(wb, layout: str) -> list[dict]:
    """抽 schedule。layout='A' 或 'B'。"""
    ws = wb['試算表']
    cols = LAYOUT_A_COLS if layout == 'A' else LAYOUT_B_COLS
    rows = []
    r = cols['data_start']

    while r <= ws.max_row:
        y = i_num(ws.cell(r, cols['y']).value)
        if y is None:
            break

        age = i_num(ws.cell(r, cols['age']).value)
        # 超過 110 歲跳過(保單終止)
        if age is not None and age > 110:
            r += 1
            continue

        cum_prem = f_num(ws.cell(r, cols['cum_prem']).value)
        cum_prem_gross = f_num(ws.cell(r, cols['cum_prem_gross']).value) \
            if 'cum_prem_gross' in cols else None
        db = f_num(ws.cell(r, cols['death_benefit']).value)
        cv = f_num(ws.cell(r, cols['cv_total']).value)
        pv = f_num(ws.cell(r, cols['policy_value']).value)
        rpu = f_num(ws.cell(r, cols['reduced_paid_up']).value)

        survival = f_num(ws.cell(r, cols['survival']).value) \
            if 'survival' in cols else None
        survival_cum = f_num(ws.cell(r, cols['survival_cum']).value) \
            if 'survival_cum' in cols else None

        # 三情境
        scenarios = {}
        for tag in ['none', 'mid', 'low']:
            div_y = f_num(ws.cell(r, cols[f'{tag}_div_year']).value)
            db_div = f_num(ws.cell(r, cols[f'{tag}_div_cum_db']).value)
            cv_div = f_num(ws.cell(r, cols[f'{tag}_div_cum_cv']).value)
            sc = {}
            if div_y is not None: sc['dividend_year'] = div_y
            if db_div is not None: sc['db_with_dividend'] = db_div
            if cv_div is not None: sc['cv_total'] = cv_div
            if sc:
                scenarios[tag] = sc

        row = {'y': y}
        if age is not None: row['age'] = age
        if cum_prem is not None: row['cum_prem'] = cum_prem
        if cum_prem_gross is not None: row['cum_prem_gross'] = cum_prem_gross
        if cv is not None:
            row['cv_basic'] = cv
            row['cv_total'] = cv
        if db is not None: row['death_benefit'] = db
        if pv is not None: row['policy_value'] = pv
        if rpu is not None and rpu != 0: row['reduced_paid_up'] = rpu
        if survival is not None and (survival != 0 or survival_cum):
            row['survival_benefit'] = survival
            if survival_cum is not None:
                row['survival_cum'] = survival_cum
        if scenarios:
            row['scenarios'] = scenarios

        # v3.1 Step 2:cv_total 跟 death_benefit 都缺,該列是空白殘留,跳過
        if 'cv_total' not in row and 'death_benefit' not in row:
            r += 1
            continue

        rows.append(row)
        r += 1

    return rows


# ============================================================================
# Step 4:自洽性驗證(13 條)
# ============================================================================

def verify(meta: dict, sched: list[dict]) -> tuple[list[str], list[str]]:
    """回傳 (errors, warnings)。errors 非空就不交付。"""
    errors = []
    warnings = []
    p = meta['period']

    if not sched:
        errors.append("schedule 是空的")
        return errors, warnings

    y1 = sched[0]
    sa = meta['base_sa']

    # 1. Y1 cum_prem ≈ base_premium
    if 'cum_prem' in y1 and meta.get('base_premium'):
        diff = abs(y1['cum_prem'] - meta['base_premium'])
        if diff > 1:
            errors.append(f"#1 Y1 cum_prem ({y1['cum_prem']}) ≠ base_premium ({meta['base_premium']}) | diff={diff:.2f}")

    # 2. Y(period) cum_prem ≈ base_premium × period
    if len(sched) >= p and meta.get('base_premium'):
        expected = meta['base_premium'] * p
        diff = abs(sched[p-1]['cum_prem'] - expected)
        if diff > p:
            errors.append(f"#2 Y{p} cum_prem ({sched[p-1]['cum_prem']}) ≠ {expected:.0f} | diff={diff:.2f}")

    # 3. Y(period+1) cum_prem 不再增加
    if len(sched) > p:
        diff = abs(sched[p]['cum_prem'] - sched[p-1]['cum_prem'])
        if diff > 1:
            errors.append(f"#3 Y{p+1} cum_prem ({sched[p]['cum_prem']}) ≠ Y{p} ({sched[p-1]['cum_prem']})")

    # 4. cv_total >= cv_basic
    for r in sched:
        if r.get('cv_total', 0) < r.get('cv_basic', 0) - 1:
            errors.append(f"#4 Y{r['y']} cv_total < cv_basic"); break

    # 5. cv_total 中後期遞增(對還本型放寬:容忍每年微降 < 2% 因為領回生存金)
    is_payout = any(r.get('survival_benefit', 0) for r in sched)
    tolerance_pct = 0.02 if is_payout else 0.001
    for i in range(10, len(sched)):
        prev_cv = sched[i-1].get('cv_total', 0)
        curr_cv = sched[i].get('cv_total', 0)
        if curr_cv < prev_cv * (1 - tolerance_pct):
            errors.append(f"#5 Y{sched[i]['y']} cv_total {curr_cv} 下降超過容忍 {tolerance_pct*100:.1f}%")
            break

    # 6. ⭐ v3.1 寬鬆:db ≈ base_sa,或身故倍增型(末 5 年平均在 [0.5, 25])
    matched = next((r for r in sched if abs(r.get('death_benefit', 0) - sa) / sa < 0.05), None)
    if matched is None:
        ratios = [r.get('death_benefit', 0) / sa for r in sched]
        max_ratio = max(ratios) if ratios else 0
        last_5_avg = sum(ratios[-5:]) / 5 if len(ratios) >= 5 else (ratios[0] if ratios else 0)
        if 0.5 <= last_5_avg <= 25:
            warnings.append(f"#6 db ≠ base_sa,身故倍增型(Y1={ratios[0]:.1f}x, max={max_ratio:.1f}x, 末5年均={last_5_avg:.1f}x)")
        else:
            errors.append(f"#6 db 比例異常:max={max_ratio:.2f}, 末5年均={last_5_avg:.2f}")

    # 7. 末年齡 ≤ 110
    last_age = sched[-1].get('age', meta.get('base_age', 0) + sched[-1]['y'] - 1)
    if last_age > 110:
        errors.append(f"#7 末年齡 {last_age} > 110")

    # 8. 筆數 >= 50(警告)
    if len(sched) < 50:
        warnings.append(f"#8 schedule 只有 {len(sched)} 筆")

    # 9. discount 自洽:gross × (1 - discount) ≈ disc
    if (meta.get('base_premium_gross') and meta.get('base_premium')
            and meta.get('discount', 0) > 0):
        gross = meta['base_premium_gross']
        net = meta['base_premium']
        disc = meta['discount']
        expected = gross * (1 - disc)
        if abs(expected - net) > 1:
            errors.append(f"#9 discount 不自洽:{gross}×(1-{disc})={expected:.2f},實 {net}")

    # 10. prudential_v2 三情境結構完整
    if meta['engine'] == 'prudential_v2':
        for r in sched:
            sc = r.get('scenarios', {})
            for name in ['none', 'mid', 'low']:
                if name not in sc:
                    errors.append(f"#10 Y{r['y']} 缺 scenarios.{name}")
                    break
                for k in ['dividend_year', 'db_with_dividend', 'cv_total']:
                    if k not in sc[name]:
                        errors.append(f"#10 Y{r['y']} scenarios.{name}.{k} 缺")
                        break

    # 11. ⭐ v3.1 寬鬆:整個 schedule 至少有一年 mid.dividend_year != 0
    if meta['engine'] == 'prudential_v2':
        any_mid_div = any(r.get('scenarios', {}).get('mid', {}).get('dividend_year', 0)
                          for r in sched)
        if not any_mid_div:
            errors.append(f"#11 整個 schedule mid.dividend_year 全為 0(分紅未生效)")

    # 12. mid.db_with_dividend >= death_benefit
    if meta['engine'] == 'prudential_v2':
        for r in sched:
            mid_db = r.get('scenarios', {}).get('mid', {}).get('db_with_dividend', 0)
            if mid_db and mid_db < r.get('death_benefit', 0) - 1:
                errors.append(f"#12 Y{r['y']} mid.db_with_dividend < death_benefit")
                break

    # 13. 末年三情境 cv:mid >= low >= none(或至少 mid >= none)
    if meta['engine'] == 'prudential_v2' and sched:
        last_sc = sched[-1].get('scenarios', {})
        if all(k in last_sc for k in ['none', 'mid', 'low']):
            mid_cv = last_sc['mid'].get('cv_total', 0)
            none_cv = last_sc['none'].get('cv_total', 0)
            if mid_cv < none_cv - 1:
                errors.append(f"#13 末年 mid.cv ({mid_cv}) < none.cv ({none_cv})")

    return errors, warnings


# ============================================================================
# 主處理流程
# ============================================================================

@dataclass
class ProcessResult:
    code: str
    layout: Optional[str] = None
    json_path: Optional[Path] = None
    skip_reason: Optional[str] = None
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    base: Optional[BaseParams] = None
    schedule_len: int = 0


def process_file(xls_path: Path, output_dir: Path,
                 force_overwrite: bool = False,
                 skip_validation: bool = False,
                 check_only: bool = False,
                 verbose: bool = False) -> ProcessResult:
    """處理單檔。"""
    fname = xls_path.name
    plan_code = extract_plan_code(fname) or '?'

    print(f"\n{'='*80}")
    print(f"處理: {fname}")
    print(f"檔名代號: {plan_code}")
    print(f"{'='*80}")

    result = ProcessResult(code=plan_code)

    # 轉檔
    try:
        if xls_path.suffix.lower() == '.xls':
            xlsx_path = convert_xls_to_xlsx(xls_path, verbose=verbose)
            print(f"  ✅ 已轉換為 .xlsx")
        else:
            xlsx_path = xls_path
    except Exception as e:
        result.errors.append(f"轉檔失敗: {e}")
        return result

    # 開檔
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
    except Exception as e:
        result.errors.append(f"開檔失敗: {e}")
        return result

    # === Step 0.0:檔名 vs 內容 ===
    print(f"\n[Step 0.0] 檔名 vs Excel 內容驗證")
    fn_check = check_filename_vs_content(xlsx_path, wb)
    print(fn_check.report())
    if not fn_check.matched and fn_check.has_actual:
        result.skip_reason = (
            f"檔名錯誤:檔名 {fn_check.expected_code},Excel 是 "
            f"{fn_check.actual_code}({fn_check.actual_product_name})"
        )
        wb.close()
        return result

    # 用 Excel 內代號取代檔名(以防檔名 regex 抽錯)
    if fn_check.actual_code:
        plan_code = fn_check.actual_code
        result.code = plan_code

    # === Step 0.5:重複版本 ===
    print(f"\n[Step 0.5] 重複版本檢測")
    ver_check = check_version_overwrite(plan_code, fname, output_dir)
    print(ver_check.report())
    if (ver_check.has_existing and not ver_check.should_overwrite
            and not force_overwrite):
        result.skip_reason = (
            f"已有 {plan_code}.json (日期 {ver_check.existing_date}),"
            f"新檔日期 {ver_check.new_date} 不較新"
        )
        wb.close()
        return result

    # === Step 0.6:layout 偵測 ===
    print(f"\n[Step 0.6] Layout 偵測")
    layout = detect_layout(wb)
    print(f"layout = {layout}")
    if layout is None:
        result.skip_reason = "不是 prudential_v2 結構(無試算表 sheet 或結構不識別)"
        wb.close()
        return result
    result.layout = layout

    if check_only:
        print("\n--check-only 模式,只驗證不抽")
        wb.close()
        return result

    # === Step 1:base 抽取 ===
    print(f"\n[Step 1] 基準參數抽取")
    base = extract_base_params(wb)
    result.base = base

    missing = base.required_filled()
    if missing:
        result.errors.append(f"必抽 base 參數缺失: {missing}")
        wb.close()
        return result

    print(f"  sex={base.sex}, age={base.age}, sa={base.sa}, period={base.period}")
    print(f"  gross={base.gross_premium}, disc={base.disc_premium}, "
          f"discount={base.discount*100:.2f}%")
    if base.payout_period is not None:
        print(f"  payout_period={base.payout_period}")

    # === Step 2:schedule 抽取 ===
    print(f"\n[Step 2] Schedule 抽取(layout {layout})")
    schedule = extract_schedule(wb, layout)
    result.schedule_len = len(schedule)
    print(f"  抽到 {len(schedule)} 筆")
    if schedule:
        print(f"  Y1: cum_prem={schedule[0].get('cum_prem')}, "
              f"db={schedule[0].get('death_benefit')}, "
              f"cv={schedule[0].get('cv_total')}")

    # === 組裝 meta ===
    pname = base.product_name or ''
    if pname.startswith(COMPANY):
        pname_short = pname[len(COMPANY):]
    else:
        pname_short = pname

    meta = {
        'product_id': plan_code,
        'company': COMPANY,
        'product_name': pname_short,
        'currency': detect_currency(pname),
        'period': base.period,
        'engine': ENGINE,
        'base_sex': base.sex,
        'base_age': base.age,
        'base_sa': base.sa,
        'base_premium': base.disc_premium,
        'base_premium_gross': base.gross_premium,
        'discount': base.discount,
        'declared_rate': 0,
        'source_file': fname,
        'extracted_at': date.today().isoformat(),
    }
    if base.discount_label is not None:
        meta['discount_label'] = base.discount_label
    if base.payout_period is not None:
        meta['payout_period'] = base.payout_period

    out_data = {'meta': meta, 'schedule': schedule}

    # === Step 4:驗證 ===
    if not skip_validation:
        print(f"\n[Step 4] 自洽性驗證")
        errors, warnings = verify(meta, schedule)
        result.errors = errors
        result.warnings = warnings

        for e in errors:
            print(f"  ❌ {e}")
        for w in warnings:
            print(f"  ⚠️  {w}")

        if errors:
            print(f"\n❌ 驗證失敗,不寫入 JSON")
            wb.close()
            return result

        print(f"  ✅ 通過(13/13)" if not warnings else
              f"  ✅ 通過(含 {len(warnings)} 項警告)")

    # === Step 5:寫 JSON ===
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f'{plan_code}.json'
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(out_data, f, ensure_ascii=False, indent=2)

    result.json_path = out_path
    print(f"\n✅ 寫出: {out_path}")

    wb.close()
    return result


# ============================================================================
# CLI
# ============================================================================

def expand_paths(paths: list[str]) -> list[Path]:
    """展開 glob,支援多個檔案。"""
    expanded = []
    for p in paths:
        if '*' in p or '?' in p:
            expanded.extend(Path(x) for x in sorted(glob.glob(p)))
        else:
            expanded.append(Path(p))
    return [p for p in expanded if p.exists()]


def main():
    parser = argparse.ArgumentParser(
        description=f'保誠 Excel → JSON 抽取器 v{VERSION}',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument('files', nargs='+', help='輸入 .xls 或 .xlsx 檔案(可多個 / glob)')
    parser.add_argument('--output', '-o', default=DEFAULT_OUTPUT_DIR,
                        help=f'輸出目錄(預設 {DEFAULT_OUTPUT_DIR})')
    parser.add_argument('--batch', action='store_true',
                        help='批次模式:單檔失敗繼續處理下一個')
    parser.add_argument('--check-only', action='store_true',
                        help='只跑 Step 0(檔名 vs 內容、版本、layout),不抽資料')
    parser.add_argument('--force-overwrite', action='store_true',
                        help='強制覆蓋現有 JSON,不檢查版本日期')
    parser.add_argument('--skip-validation', action='store_true',
                        help='跳過自洽性驗證(慎用)')
    parser.add_argument('-v', '--verbose', action='store_true', help='詳細輸出')
    parser.add_argument('--version', action='version', version=f'%(prog)s {VERSION}')

    args = parser.parse_args()

    paths = expand_paths(args.files)
    if not paths:
        print(f"❌ 沒有找到任何檔案", file=sys.stderr)
        sys.exit(4)

    output_dir = Path(args.output)

    print(f"輸入檔案數: {len(paths)}")
    print(f"輸出目錄: {output_dir}")
    print(f"模式: {'批次' if args.batch else '單檔'}"
          f"{' / 只驗證' if args.check_only else ''}"
          f"{' / 強制覆蓋' if args.force_overwrite else ''}"
          f"{' / 跳過驗證' if args.skip_validation else ''}")

    results = []
    for path in paths:
        try:
            r = process_file(
                path, output_dir,
                force_overwrite=args.force_overwrite,
                skip_validation=args.skip_validation,
                check_only=args.check_only,
                verbose=args.verbose,
            )
            results.append(r)
        except Exception as e:
            r = ProcessResult(code='?', errors=[f"系統錯誤: {e}"])
            results.append(r)
            if args.verbose:
                import traceback
                traceback.print_exc()
            if not args.batch:
                print(f"\n❌ 系統錯誤: {e}", file=sys.stderr)
                sys.exit(4)

    # === 總結 ===
    print(f"\n{'='*80}")
    print(f"總結 ({len(results)} 個檔案)")
    print(f"{'='*80}")

    success = [r for r in results if r.json_path]
    skipped = [r for r in results if r.skip_reason]
    failed = [r for r in results if r.errors and not r.json_path and not r.skip_reason]

    print(f"\n✅ 成功 ({len(success)} 個):")
    for r in success:
        warnings_note = f" ({len(r.warnings)} warn)" if r.warnings else ""
        print(f"  {r.code} [layout {r.layout}] → {r.json_path.name}{warnings_note}")

    if skipped:
        print(f"\n⏭️  跳過 ({len(skipped)} 個):")
        for r in skipped:
            print(f"  {r.code}: {r.skip_reason}")

    if failed:
        print(f"\n❌ 失敗 ({len(failed)} 個):")
        for r in failed:
            print(f"  {r.code}:")
            for e in r.errors[:3]:
                print(f"    - {e}")

    # 退出碼
    if failed:
        sys.exit(1)
    if not success and skipped:
        sys.exit(2)
    if not success:
        sys.exit(3)
    sys.exit(0)


if __name__ == '__main__':
    main()
