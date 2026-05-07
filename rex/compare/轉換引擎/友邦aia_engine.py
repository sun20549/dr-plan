"""
友邦人壽商品 Excel → JSON 抽取引擎 v1.0
================================================
依據 v4 萬用整合版指令 + 23 商品實戰經驗整合。

單檔用法：
    from aia_engine import AIAExtractor
    ext = AIAExtractor('UWHL-V2.xlsx', plan_code='UWHL-V2')
    result = ext.extract()
    if result:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print(ext.errors)

批次用法請用 aia_extract_cli.py。

已驗證商品（23 個）：
    UED, UWHL, NHISWL, NISL, NISWL, UDISRI, UDISWL, UFISWL, UHISWL,
    UISWL, URO, USWLB, USWLE, USWLF, USWLG, USWLH, UWLS, UWLV,
    BMW7PIS2, NWLCA, NWLS, NWLV, UWLC

已知不支援：
    NRI（5599 還本，PREM 欄位特殊）
    NISDA（年金險，結構完全不同）
"""

import openpyxl
import json
import datetime
import os
import re
from pathlib import Path
from typing import Optional, Dict, Any, List, Tuple


__version__ = '1.0'
__all__ = ['AIAExtractor']


# ============ 常數 ============

CURRENCY_DEFAULTS = {
    'USD': {'min_sa': 10000, 'max_sa': 5000000, 'unit_size': 1000, 'sa_per_unit_div': 1000,
            'mid_div': 0.055},
    'TWD': {'min_sa': 300000, 'max_sa': 100000000, 'unit_size': 1000, 'sa_per_unit_div': 10000,
            'mid_div': 0.045},
    'AUD': {'min_sa': 10000, 'max_sa': 3000000, 'unit_size': 10000, 'sa_per_unit_div': 1000,
            'mid_div': 0.050},
    'CNY': {'min_sa': 50000, 'max_sa': 30000000, 'unit_size': 1000, 'sa_per_unit_div': 10000,
            'mid_div': 0.040},
}

# type 字串對應表（Step 6 規則 B）
def infer_type_str(currency: str, has_dividend: bool, has_survival: bool,
                    is_endowment: bool) -> str:
    cur_zh = {'USD': '美元', 'TWD': '新台幣', 'AUD': '澳幣', 'CNY': '人民幣'}.get(currency, '未知')
    suffix = ''
    if has_dividend:
        suffix = '分紅'
    else:
        suffix = '利率變動型'
    body = ''
    if is_endowment:
        body = '養老保險'
    elif has_survival:
        body = '還本終身壽險' if has_dividend else '還本終身壽險'
    else:
        body = '終身壽險'
    return f'{cur_zh}{suffix}{body}'


# ============ 工具 ============

def to_num(v):
    """處理 LibreOffice 轉檔後字串 / None / '-' 等。v4 F0.4 規範。"""
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        s = v.strip().replace(',', '').replace('$', '').replace(' ', '')
        if s in ('-', '－', '—', 'N/A', '#N/A', '#VALUE!', '#REF!', '#DIV/0!'):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def is_error_string(v) -> bool:
    """偵測 #VALUE! / #REF! 等公式爆掉標記"""
    if not isinstance(v, str):
        return False
    return any(tag in v for tag in ('#VALUE', '#REF', '#DIV', '#NAME', '#N/A'))


# ============ 主引擎 ============

class AIAExtractor:
    """
    友邦人壽商品 RV 表手算引擎。
    
    流程:
        Step 0: 結構偵察(self._inspect_structure)
        Step 0.5: 公式爆掉處理(由外部 aia_recalc.py 預先處理)
        Step 1: 從輸入頁抽 base 參數(self._scan_input_page)
        Step 2: 從 FACTOR/PREM 表查資料 + 手算 schedule(self._build_schedule)
        Step 3: 組 JSON
        Step 4: 自洽性驗證(self._verify)
        Step 5: 回傳 dict
    """
    
    # 友邦商品 plan_short 已知別名（避免每次 fallback 都警告）
    KNOWN_MULTI_PLAN_CD = {
        'UDISRI', 'UWLS', 'UWLV', 'NWLS', 'NWLV', 'NWLCA', 'URO', 'NRI',
        'UWLC',  # 嬰兒投保多 D1?11T 起始日方案
    }
    
    # 已知不支援的 plan_short
    UNSUPPORTED_PLAN_SHORT = {
        'NRI',     # 5599 還本,PREM 表 PPP 欄存 plan_cd 編碼
        'NISDA',   # 年金險,無 FACTOR/PREM 結構
    }
    
    def __init__(self, xlsx_path: str, plan_code: str = None):
        self.path = Path(xlsx_path)
        # 從檔名抽 plan_code (例 _UWHL-V2_ → UWHL-V2)
        if plan_code is None:
            m = re.search(r'_([A-Z][A-Z0-9]+(?:-V\d+)?)_', self.path.name)
            plan_code = m.group(1) if m else self.path.stem
        self.plan_code = plan_code
        self.plan_short = plan_code.split('-')[0]
        
        if not self.path.exists():
            raise FileNotFoundError(f"檔案不存在: {xlsx_path}")
        
        if self.plan_short in self.UNSUPPORTED_PLAN_SHORT:
            raise NotImplementedError(
                f"plan_short={self.plan_short} 已知不支援。"
                f"NRI 結構特殊（PREM PPP 欄存編碼），NISDA 為年金險（架構不同）。"
            )
        
        self.wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        self.warnings: List[str] = []
        self.errors: List[str] = []
    
    # ---------- Step 0: 結構偵察 ----------
    
    def _find_sheet(self, candidates: List[str]):
        """大小寫不敏感找 sheet"""
        names_lower = {s.lower(): s for s in self.wb.sheetnames}
        for c in candidates:
            if c.lower() in names_lower:
                return self.wb[names_lower[c.lower()]]
        return None
    
    def _inspect_structure(self) -> Dict[str, Any]:
        """偵察 sheet 結構。回傳 {has_factor, has_prem, has_input, ...}"""
        sheets = self.wb.sheetnames
        info = {
            'sheets': sheets,
            'has_factor': any(s.lower() == 'factor' for s in sheets),
            'has_prem': any(s.lower() in ('prem', 'premium') for s in sheets),
            'has_input': '輸入頁' in sheets,
            'has_analysis': '保險利益分析表' in sheets,
        }
        info['is_aia_rv'] = (info['has_factor'] and info['has_prem'] 
                             and info['has_input'] and info['has_analysis'])
        return info
    
    # ---------- Step 1: 抽輸入頁參數 ----------
    
    def _scan_input_page(self) -> Dict[str, Any]:
        ws = self._find_sheet(['輸入頁'])
        if ws is None:
            self.errors.append("找不到輸入頁")
            return {}
        
        # 重新打開非 read_only 模式，方便 cell access
        wb_rw = openpyxl.load_workbook(self.path, data_only=True)
        ws_rw = wb_rw['輸入頁']
        max_r = min(ws_rw.max_row or 100, 100)
        max_c = min(ws_rw.max_column or 50, 50)
        cells: Dict[Tuple[int, int], Any] = {}
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                v = ws_rw.cell(r, c).value
                if v is not None:
                    cells[(r, c)] = v
        wb_rw.close()
        
        def find_label(keyword: str):
            """產生所有含 keyword 的 (r, c, value) tuple"""
            for (r, c), v in cells.items():
                if isinstance(v, str) and keyword in v:
                    yield (r, c, v)
        
        def value_after(label_pos: Tuple[int, int], max_dc: int = 8,
                        want_type: Optional[str] = None):
            """從 label 位置往右找第一個合用值"""
            r, c = label_pos
            for dc in range(1, max_dc + 1):
                target = cells.get((r, c + dc))
                if target is None:
                    continue
                if is_error_string(target):
                    continue
                if want_type == 'num' and not isinstance(target, (int, float)):
                    continue
                if want_type == 'str' and not isinstance(target, str):
                    continue
                return target, (r, c + dc)
            return None, None
        
        params: Dict[str, Any] = {}
        
        # === 性別 ===
        for r, c, v in find_label('性別'):
            sex_v, _ = value_after((r, c), want_type='str')
            if sex_v and ('男' in sex_v or '女' in sex_v):
                params['base_sex'] = 'M' if '男' in sex_v else 'F'
                break
        
        # === 生日(民國年 + 月)→ 推算保險年齡 ===
        # 友邦規則:115 - 生日年,生日月 > 1 要 -1
        for r, c, v in find_label('生日'):
            year_v, year_pos = value_after((r, c), want_type='num')
            if isinstance(year_v, (int, float)) and 0 <= year_v <= 130:
                params['_birth_year'] = int(year_v)
                if year_pos:
                    for dc in range(1, 8):
                        m = cells.get((year_pos[0], year_pos[1] + dc))
                        if isinstance(m, (int, float)) and 1 <= m <= 12:
                            params['_birth_month'] = int(m)
                            break
                break
        
        if '_birth_year' in params:
            birth_year = params['_birth_year']
            birth_month = params.get('_birth_month', 1)
            age = 115 - birth_year  # 民國 115 年(2026)為基準
            if birth_month > 1:
                age -= 1  # 還沒到今年生日
            # base_age 範圍:0(嬰兒投保) ~ 80
            if 0 <= age <= 80:
                params['base_age'] = age
        
        # === 繳費年期 ===
        for r, c, v in find_label('繳費年期'):
            period_v, _ = value_after((r, c), want_type='num')
            if isinstance(period_v, (int, float)) and 1 <= period_v <= 30:
                params['period'] = int(period_v)
                break
        
        # === 投保金額 / 保險金額 / 單張保額 ===
        for keyword in ['投保金額', '保險金額', '單張保額']:
            found = False
            for r, c, v in find_label(keyword):
                sa_v, sa_pos = value_after((r, c), want_type='num')
                if isinstance(sa_v, (int, float)) and sa_v >= 1:
                    sa_unit = None
                    if sa_pos:
                        for dc in range(1, 6):
                            u = cells.get((sa_pos[0], sa_pos[1] + dc))
                            if isinstance(u, str) and any(unit in u for unit in 
                                ('萬元', '萬美元', '千美元', '千元', '美元', '萬', '元')):
                                sa_unit = u
                                break
                    if sa_unit:
                        params['_sa_raw'] = sa_v
                        params['_sa_unit'] = sa_unit
                        found = True
                        break
            if found:
                break
        
        # === 假設宣告利率 ===
        for r, c, v in find_label('假設宣告利率'):
            rate_v, _ = value_after((r, c), want_type='num')
            if isinstance(rate_v, (int, float)) and 0 < rate_v < 1:
                params['declared_rate'] = round(float(rate_v), 4)
                break
        
        # === 幣別 ===
        for (r, c), v in cells.items():
            if isinstance(v, str) and '幣別' in v:
                if '美元' in v:
                    params['currency'] = 'USD'
                elif '澳幣' in v or '澳元' in v:
                    params['currency'] = 'AUD'
                elif '人民幣' in v:
                    params['currency'] = 'CNY'
                elif '新台幣' in v or '台幣' in v:
                    params['currency'] = 'TWD'
                break
        
        # 沒寫幣別 → 從商品名推
        if 'currency' not in params:
            for (r, c), v in cells.items():
                if isinstance(v, str) and '建議書' in v:
                    if '美元' in v:
                        params['currency'] = 'USD'
                    elif '澳' in v:
                        params['currency'] = 'AUD'
                    elif '人民幣' in v:
                        params['currency'] = 'CNY'
                    else:
                        params['currency'] = 'TWD'
                    break
        
        # === 商品全名 ===
        for (r, c), v in cells.items():
            if isinstance(v, str) and '建議書' in v:
                name = v.replace('--建議書', '').strip()
                params['product_name'] = name
                break
        
        # === 預定利率(從字串中 regex 抽)===
        for (r, c), v in cells.items():
            if isinstance(v, str) and '預定利率' in v:
                m = re.search(r'(\d+(?:\.\d+)?)\s*%', v)
                if m:
                    params['guaranteed_rate'] = round(float(m.group(1)) / 100, 4)
                    break
        
        # === 還本商品偵測 ===
        is_endow = False
        endow_signal = None
        for (r, c), v in cells.items():
            if isinstance(v, str):
                for kw in ('生存保險金', '教育基金', '還本'):
                    if kw in v:
                        is_endow = True
                        endow_signal = kw
                        break
                if is_endow:
                    break
        # 也從商品名判斷
        if not is_endow and params.get('product_name'):
            if '還本' in params['product_name'] or '養老' in params['product_name']:
                is_endow = True
                endow_signal = '商品名'
        params['_is_endowment'] = is_endow
        params['_endow_signal'] = endow_signal
        
        # === 養老型偵測(商品名含「養老」)===
        if params.get('product_name') and '養老' in params['product_name']:
            params['_is_endowment_type'] = True
        
        # === base_sa ===
        if params.get('_sa_raw') is not None and params.get('_sa_unit'):
            sa = params['_sa_raw']
            unit = params['_sa_unit']
            if '萬美元' in unit:
                params['base_sa'] = int(sa * 10000)
            elif '千美元' in unit:
                params['base_sa'] = int(sa * 1000)
            elif '萬元' in unit or '萬' in unit:
                params['base_sa'] = int(sa * 10000)
            elif '千元' in unit or '千' in unit:
                params['base_sa'] = int(sa * 1000)
            else:
                params['base_sa'] = int(sa)
        
        # === 折扣表 ===
        discount_table = []
        for (r, c), v in cells.items():
            if isinstance(v, str) and '元' in v and ('保額' in v or '≦' in v) and '折扣' not in v:
                for dc in range(1, 8):
                    rate_v = cells.get((r, c + dc))
                    if isinstance(rate_v, (int, float)) and 0 < rate_v < 0.5:
                        discount_table.append((v, rate_v))
                        break
        params['_discount_table'] = discount_table
        
        # === declared_rate fallback(有些純保證型還本/養老沒這欄)===
        if 'declared_rate' not in params:
            if 'guaranteed_rate' in params:
                params['declared_rate'] = params['guaranteed_rate']
                self.warnings.append(
                    f"無假設宣告利率,fallback 用預定利率 {params['guaranteed_rate']}"
                )
            else:
                params['declared_rate'] = 0
                self.warnings.append("declared_rate=0(可能純保證型)")
        
        return params
    
    # ---------- Step 2: 查 FACTOR/PREM ----------
    
    def _detect_table_schema(self, ws) -> Dict[str, int]:
        """偵測 FACTOR/PREM 表的 col 含義(動態 schema)"""
        try:
            header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        except (StopIteration, IndexError):
            return {}
        
        schema: Dict[str, int] = {}
        # 用 setdefault 避免重複欄位被後面覆蓋(UDISRI Premium 有兩個 plan_cd 欄)
        for i, h in enumerate(header):
            if h is None:
                continue
            hl = str(h).strip().lower()
            if hl in ('plan', 'plan_kind'):
                schema.setdefault('plan', i)
            elif hl in ('plan_cd', 'plan code', 'plan_code'):
                schema.setdefault('plan_cd', i)
            elif hl == 'ppp':
                schema.setdefault('ppp', i)
            elif hl == 'age':
                schema.setdefault('age', i)
            elif hl in ('sex', 'gender'):
                schema.setdefault('sex', i)
            elif hl == 'factor':
                schema.setdefault('factor', i)
            elif hl == 'prem':
                schema.setdefault('prem', i)
            elif hl.startswith('dur'):
                schema.setdefault('dur_start', i)
        return schema
    
    def _do_query(self, factor_ws, prem_ws, plan_key: str,
                  base_age: int, base_sex: str, period: int):
        """單次查詢 FACTOR + PREM"""
        f_sch = self._detect_table_schema(factor_ws)
        if 'age' not in f_sch or 'sex' not in f_sch or 'factor' not in f_sch:
            return None, None
        
        def factor_match(row):
            try:
                plan_ok = False
                if 'plan' in f_sch and str(row[f_sch['plan']]) == plan_key:
                    plan_ok = True
                if not plan_ok and 'plan_cd' in f_sch and str(row[f_sch['plan_cd']]) == plan_key:
                    plan_ok = True
                if not plan_ok and 'plan' not in f_sch and 'plan_cd' not in f_sch:
                    plan_ok = True
                if not plan_ok:
                    return False
                if int(row[f_sch['age']]) != base_age:
                    return False
                if str(row[f_sch['sex']]) != base_sex:
                    return False
                if 'ppp' in f_sch and int(row[f_sch['ppp']]) != period:
                    return False
                return True
            except (TypeError, ValueError, IndexError):
                return False
        
        # 抽 FACTOR
        factors: Dict[str, List] = {}
        dur_start = f_sch.get('dur_start')
        if dur_start is None:
            return None, None
        for row in factor_ws.iter_rows(min_row=2, values_only=True):
            if factor_match(row):
                fname = row[f_sch['factor']]
                durs = list(row[dur_start + 1:])  # skip dur0
                # 截斷到非 None 部分
                valid_count = 0
                for v in durs:
                    if v is None:
                        break
                    valid_count += 1
                factors[fname] = durs[:valid_count]
        
        if not factors or 'CV' not in factors or 'DB' not in factors:
            return None, None
        
        # PREM
        prem_value = None
        if prem_ws is not None:
            p_sch = self._detect_table_schema(prem_ws)
            if 'age' in p_sch and 'sex' in p_sch and 'prem' in p_sch:
                def prem_match(row):
                    try:
                        plan_ok = False
                        if 'plan' in p_sch and str(row[p_sch['plan']]) == plan_key:
                            plan_ok = True
                        if not plan_ok and 'plan_cd' in p_sch and str(row[p_sch['plan_cd']]) == plan_key:
                            plan_ok = True
                        if not plan_ok and 'plan' not in p_sch and 'plan_cd' not in p_sch:
                            plan_ok = True
                        if not plan_ok:
                            return False
                        if int(row[p_sch['age']]) != base_age:
                            return False
                        if str(row[p_sch['sex']]) != base_sex:
                            return False
                        if 'ppp' in p_sch and int(row[p_sch['ppp']]) != period:
                            return False
                        return True
                    except (TypeError, ValueError, IndexError):
                        return False
                
                for row in prem_ws.iter_rows(min_row=2, values_only=True):
                    if prem_match(row):
                        prem_value = row[p_sch['prem']]
                        break
        
        if prem_value is None:
            return None, None
        
        return prem_value, factors
    
    def _query_factor_prem(self, base_age: int, base_sex: str, period: int):
        """主查詢入口,含多 plan_cd fallback"""
        prem_ws = self._find_sheet(['PREM', 'Prem', 'Premium'])
        factor_ws = self._find_sheet(['FACTOR', 'Factor'])
        
        if factor_ws is None:
            self.errors.append("找不到 FACTOR sheet")
            return None, None, None
        
        # 第 1 輪:用 plan_short 直查
        prem, factors = self._do_query(factor_ws, prem_ws, self.plan_short,
                                         base_age, base_sex, period)
        if prem is not None:
            return prem, factors, period
        
        # 第 2 輪:多 plan_cd fallback
        f_sch = self._detect_table_schema(factor_ws)
        plan_col = f_sch.get('plan_cd', f_sch.get('plan'))
        if plan_col is not None:
            candidates = set()
            for row in factor_ws.iter_rows(min_row=2, values_only=True):
                try:
                    if (int(row[f_sch['age']]) == base_age and
                            str(row[f_sch['sex']]) == base_sex and
                            ('ppp' not in f_sch or int(row[f_sch['ppp']]) == period)):
                        candidates.add(str(row[plan_col]))
                except (TypeError, ValueError, IndexError):
                    pass
            
            if candidates:
                chosen = sorted(candidates)[0]
                self.warnings.append(
                    f"多 plan_cd fallback,取 {chosen}({len(candidates)} 候選);"
                    f"請依業務員填的方案校對 (UWLS 起始日 / UDISRI 樂退年齡 等)"
                )
                prem, factors = self._do_query(factor_ws, prem_ws, chosen,
                                                 base_age, base_sex, period)
                if prem is not None:
                    return prem, factors, period
        
        # 第 3 輪:ppp fallback(該 plan 沒這個 ppp)
        if 'ppp' in f_sch:
            available_ppps = set()
            for row in factor_ws.iter_rows(min_row=2, values_only=True):
                try:
                    if (int(row[f_sch['age']]) == base_age and
                            str(row[f_sch['sex']]) == base_sex):
                        available_ppps.add(int(row[f_sch['ppp']]))
                except (TypeError, ValueError, IndexError):
                    pass
            
            if available_ppps and period not in available_ppps:
                closest = min(available_ppps, key=lambda p: abs(p - period))
                self.warnings.append(
                    f"PPP={period} 在 FACTOR 不存在,fallback 取 PPP={closest}"
                )
                # 再 retry 用 closest
                prem, factors = self._do_query(factor_ws, prem_ws, self.plan_short,
                                                 base_age, base_sex, closest)
                if prem is not None:
                    return prem, factors, closest
        
        return None, None, None
    
    # ---------- Step 2: 構築 schedule ----------
    
    def _build_schedule(self, factors: Dict[str, List], sa_per_unit: float,
                        base_age: int, period: int, prem_per_year: float) -> List[Dict]:
        sched = []
        max_y = max(len(factors.get('CV', [])), len(factors.get('DB', [])))
        cv_arr = factors.get('CV', [])
        db_arr = factors.get('DB', [])
        
        for y in range(1, max_y + 1):
            age = base_age + y - 1
            if age > 110:  # 嚴格切到 110 歲
                break
            if y - 1 >= len(cv_arr) or y - 1 >= len(db_arr):
                break
            cv = cv_arr[y - 1]
            db = db_arr[y - 1]
            if cv is None or db is None:
                break
            
            cum_prem = round(min(y, period) * prem_per_year, 2)
            sched.append({
                'y': y,
                'age': age,
                'cum_prem': cum_prem,
                'cv_basic': round(cv * sa_per_unit, 2),
                'cv_total': round(cv * sa_per_unit, 2),  # 純保證,前端疊加紅利
                'death_benefit': round(db * sa_per_unit, 2),
            })
        
        # 強制排序
        sched.sort(key=lambda r: r['y'])
        return sched
    
    # ---------- 折扣計算 ----------
    
    def _calc_discount(self, base_sa: int, discount_table: List, currency: str):
        if not discount_table:
            return 0, None
        
        applicable = 0.0
        label_used = None
        
        for desc, rate in discount_table:
            nums = re.findall(r'[\d,]+', desc)
            if not nums:
                continue
            try:
                lo = int(nums[0].replace(',', ''))
                hi = int(nums[-1].replace(',', '')) if len(nums) >= 2 else None
                
                # 單位:萬美元 / 千美元 / 萬元 / 萬(美元商品也用)
                multiplier = 1
                if '萬美元' in desc:
                    multiplier = 10000
                elif '千美元' in desc:
                    multiplier = 1000
                elif '萬元' in desc:
                    multiplier = 10000
                elif '萬' in desc and currency == 'USD':
                    multiplier = 10000  # 美元商品「萬」也是萬美元
                
                lo *= multiplier
                if hi is not None:
                    hi *= multiplier
                
                if hi is None or lo == hi:
                    if base_sa >= lo and rate > applicable:
                        applicable = rate
                        label_used = desc
                else:
                    if lo <= base_sa < hi:
                        applicable = rate
                        label_used = desc
                        break
            except (ValueError, TypeError):
                continue
        
        return applicable, label_used
    
    # ---------- 商品分類 ----------
    
    def _classify_db_type(self, sched: List[Dict], base_sa: int) -> str:
        y1_r = sched[0]['death_benefit'] / base_sa
        y6_r = sched[5]['death_benefit'] / base_sa if len(sched) >= 6 else None
        
        if 0.95 <= y1_r <= 1.05:
            return 'standard'
        elif y1_r < 0.95 and y6_r and 0.95 <= y6_r <= 1.05:
            return 'step_up_y6'
        elif y1_r > 1.05:
            return f'multiplier_{y1_r:.2f}x'
        elif y1_r < 0.95:
            return 'unusual'  # 還本險常見
        return 'unknown'
    
    # ---------- Step 4: 自洽性驗證 ----------
    
    def _verify(self, data: Dict, gross: float):
        sched = data['schedule']
        base = data['meta']
        errors: List[str] = []
        warnings: List[str] = []
        p = base['period']
        is_endow = base.get('is_endowment', False)
        is_stepped = base.get('db_pattern') == 'stepped'
        
        # 1. Y1 cum_prem ≈ base_premium
        if abs(sched[0]['cum_prem'] - base['base_premium']) > 1:
            errors.append(f"Y1 cum_prem {sched[0]['cum_prem']} ≠ base_premium {base['base_premium']}")
        
        # 2/3. 躉繳 vs 多年期
        if p == 1:
            if len(sched) > 1 and abs(sched[1]['cum_prem'] - sched[0]['cum_prem']) > 1:
                errors.append("Y2 cum_prem 應 = Y1(躉繳)")
        else:
            if len(sched) >= p:
                expected = base['base_premium'] * p
                tol = max(p, expected * 0.05)  # 5% 容差(分紅抵繳保費型)
                if abs(sched[p - 1]['cum_prem'] - expected) > tol:
                    warnings.append(f"Y{p} cum_prem 抵繳差異")
            if len(sched) > p:
                if abs(sched[p]['cum_prem'] - sched[p - 1]['cum_prem']) > 1:
                    errors.append(f"Y{p+1} cum_prem 應停止增加")
        
        # 4. cv_total >= cv_basic
        for r in sched:
            if r['cv_total'] < r['cv_basic'] - 1:
                errors.append(f"Y{r['y']} cv_total < cv_basic")
                break
        
        # 5. 中後期遞增(還本商品放寬)
        if not is_endow:
            for i in range(min(10, len(sched)), len(sched)):
                if sched[i]['cv_total'] < sched[i - 1]['cv_total'] * 0.999:
                    warnings.append(f"Y{sched[i]['y']} cv_total 微下降")
                    break
        
        # 6. db ≈ base_sa(stepped/還本放寬)
        db_max = max(r['death_benefit'] for r in sched)
        if is_stepped or is_endow:
            if db_max < base['base_sa'] * 0.95:
                errors.append(f"db_max {db_max} 從未達 base_sa")
        else:
            any_in_range = any(0.95 <= r['death_benefit'] / base['base_sa'] <= 1.05
                               for r in sched)
            if not any_in_range:
                warnings.append("無年度 db 在 [0.95-1.05] sa")
        
        # 7. age ≤ 110
        last_age = sched[-1].get('age', base['base_age'] + sched[-1]['y'] - 1)
        if last_age > 110:
            errors.append(f"末年齡 {last_age} > 110")
        
        # 8. 筆數 >= 50(養老/還本/高齡放寬)
        if (len(sched) < 50 and not is_endow and base['base_age'] < 50
                and base.get('product_type') != 'endowment'):
            warnings.append(f"筆數 {len(sched)} < 50")
        
        # 9. discount 自洽
        if gross and base.get('discount', 0) > 0:
            net = base['base_premium']
            discount = base['discount']
            expected_net = gross * (1 - discount)
            tol = max(2, gross * 0.001)
            if abs(expected_net - net) > tol:
                errors.append(f"discount 不自洽: gross×(1-{discount}) = {expected_net:.2f}, net = {net}")
        
        return errors, warnings
    
    # ---------- 主入口 ----------
    
    def extract(self) -> Optional[Dict]:
        """執行完整抽取流程。成功回傳 dict,失敗回 None(self.errors 有訊息)"""
        # Step 0: 偵察結構
        info = self._inspect_structure()
        if not info['is_aia_rv']:
            self.errors.append(
                f"非標準友邦 RV 表結構。sheets={info['sheets'][:8]}..."
            )
            return None
        
        # Step 1: 抽 base 參數
        params = self._scan_input_page()
        if self.errors:
            return None
        
        # 必抽欄位檢查(用 is None 避免 base_age=0 / base_sex='F' 被誤判)
        required = ['base_sex', 'base_age', 'base_sa', 'period', 'currency']
        for k in required:
            if params.get(k) is None:
                self.errors.append(f"缺必要參數: {k}")
                return None
        if 'declared_rate' not in params:
            params['declared_rate'] = 0  # fallback 已在 _scan 處理過
        
        # Step 2: 查 FACTOR + PREM
        prem_per_unit, factors, actual_ppp = self._query_factor_prem(
            params['base_age'], params['base_sex'], params['period']
        )
        if prem_per_unit is None:
            self.errors.append(
                f"FACTOR/PREM 查無 (plan_short={self.plan_short}, "
                f"age={params['base_age']}, sex={params['base_sex']}, "
                f"ppp={params['period']})"
            )
            return None
        
        if actual_ppp != params['period']:
            params['period'] = actual_ppp
        
        # 單位偵測:USD→每千美元, TWD→每萬元
        cur_def = CURRENCY_DEFAULTS.get(params['currency'], CURRENCY_DEFAULTS['TWD'])
        sa_per_unit = params['base_sa'] / cur_def['sa_per_unit_div']
        
        prem_gross = round(prem_per_unit * sa_per_unit, 2)
        
        # 折扣
        discount, discount_label = self._calc_discount(
            params['base_sa'], params.get('_discount_table', []), params['currency']
        )
        prem_net = round(prem_gross * (1 - discount), 2)
        
        # Step 2: 構築 schedule
        schedule = self._build_schedule(
            factors, sa_per_unit,
            params['base_age'], params['period'], prem_net
        )
        if not schedule:
            self.errors.append("schedule 為空")
            return None
        
        # 商品分類
        db_type = self._classify_db_type(schedule, params['base_sa'])
        
        # 商品名 → 規則 A: 去公司前綴 + 規則 F: 半形括號
        product_name = params.get('product_name', f'友邦人壽 {self.plan_code}')
        if product_name.startswith('友邦人壽'):
            product_name = product_name[len('友邦人壽'):]
        product_name = (product_name
                        .replace('(', '(').replace(')', ')')
                        .replace('－', '-').replace('—', '-'))
        
        # Step 3: 組 meta
        meta: Dict[str, Any] = {
            'product_id': self.plan_short,
            'company': '友邦人壽',
            'product_name': product_name,
            'currency': params['currency'],
            'period': params['period'],
            'engine': 'twlife_v1',
            'base_sex': params['base_sex'],
            'base_age': params['base_age'],
            'base_sa': params['base_sa'],
            'base_premium': prem_net,
            'base_premium_gross': prem_gross,
            'discount': round(discount, 4),
            'declared_rate': params['declared_rate'],
            'source_file': self.path.name.replace('.xlsx', '.xls'),
            'extracted_at': datetime.date.today().isoformat(),
            'extraction_note': (
                '原 Excel 公式 #VALUE!,逐年表由 FACTOR 表手算(預定利率下純保證部分);'
                '利變紅利由前端 declared_rate 推算'
            ),
        }
        
        if discount_label:
            meta['discount_label'] = discount_label
        if 'guaranteed_rate' in params:
            meta['guaranteed_rate'] = params['guaranteed_rate']
        if params.get('_is_endowment'):
            meta['is_endowment'] = True
        if params.get('_is_endowment_type'):
            meta['product_type'] = 'endowment'
            # 養老型 protection_period = schedule 長度
            meta['protection_period'] = len(schedule)
        if db_type == 'step_up_y6':
            meta['db_pattern'] = 'stepped'
            meta['step_up_year'] = 6
        meta['_db_type_inferred'] = db_type  # 內部除錯用,輸出時可剝
        
        data = {'meta': meta, 'schedule': schedule}
        
        # Step 4: 驗證
        errors, warnings = self._verify(data, prem_gross)
        if errors:
            self.errors.extend(errors)
        self.warnings.extend(warnings)
        
        return data
    
    def to_clean_dict(self, data: Dict) -> Dict:
        """剝掉 _ 開頭的 internal 欄位,給對外輸出用"""
        clean_meta = {k: v for k, v in data['meta'].items() if not k.startswith('_')}
        return {'meta': clean_meta, 'schedule': data['schedule']}
