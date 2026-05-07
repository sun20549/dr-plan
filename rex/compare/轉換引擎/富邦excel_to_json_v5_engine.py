#!/usr/bin/env python3
"""
==============================================================================
富邦/凱基/友邦/台壽/宏泰  Excel → JSON  轉換引擎  v5.0
==============================================================================

整合 v2.1 / v2.2 / v4 / v5 全部實戰經驗。
對應指令:excel_to_json_轉換指令_v5_終極版.md

支援的引擎:
  - twlife_v1        (主流利變/還本/增額型)
  - prudential_v2    (三情境分紅型)
  - simple_print_v1  (養老/還本特殊結構)

使用方式:
  from excel_to_json_v5 import process_one_file, process_batch
  
  # 單檔
  data, errors, warnings = process_one_file('xxx.xlsx', 'PFA', '富邦人壽', 
                                              '美富紅運外幣分紅終身壽險', 'auto')
  
  # 批次(配合 task list)
  results = process_batch(TASKS)

核心改進(v5 vs v4):
  ✅ currency 鎖定 R44+ 計價幣別標籤(避免法規條文誤觸發)
  ✅ base_sa 鎖定『主約險種代號』下方的『保額(萬)』(避免 col 11 註解誤抓)
  ✅ 躉繳商品獨立 keyword + 自動 fallback
  ✅ 歲滿期型偵測(period_type=age_based)
  ✅ Y1 vs Y2 差異檢查(附約保費分離 SKIP)
  ✅ 混合型偵測(意外+醫療+生存)
  ✅ warning 分級(red/yellow/green)
  ✅ currency × base_sa 數量級交叉驗證
"""

import openpyxl
import json
import re
from pathlib import Path
from datetime import date

# =============================================================================
# 共用工具
# =============================================================================

def safe_get(row, idx):
    """安全讀 row 的第 idx 個欄位"""
    return row[idx] if idx < len(row) else None


def find_value_after(row, label_idx, max_offset=8, value_filter=None):
    """從標籤位置往右找下一個非空且符合 filter 的值"""
    for j in range(label_idx + 1, min(label_idx + max_offset + 1, len(row))):
        v = row[j]
        if v is None or v == '': continue
        if value_filter is None or value_filter(v):
            return v, j
    return None, None


def extract_continuous_y(ws, data_start_row, y_col=1):
    """抽連續 Y 序列,Y reset 即停"""
    rows = []
    prev_y = 0
    for r in range(data_start_row, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        y = safe_get(row, y_col)
        if not (isinstance(y, int) and 1 <= y <= 110): continue
        age = safe_get(row, 2)
        if not (isinstance(age, int) and age <= 110): continue
        if y < prev_y: break
        rows.append((r, row))
        prev_y = y
    return rows


# =============================================================================
# Step 0:結構偵察 + 類型判斷
# =============================================================================

SKIP_KEYWORDS = {
    '變額/投資型': ['變額', '萬能', '投資型', 'UL', 'Universal', 'Variable', '假設投資報酬率'],
    '年金險':     ['即期年金', '變額年金', '遞延年金'],
    '醫療險':     ['醫療', '醫保', '醫卡', '健康保險', '住院'],
    '防癌險':     ['防癌', '癌無憂', '愛無憂', '癌症', '無憂'],
    '重疾險':     ['重大疾病', '重大傷病', '心關懷', '繡情', '丰彩'],
    '長照險':     ['長期照顧', '長照', '失能', '扶照', '扶保', '照護'],
    '意外/平安':   ['傷害保險', '意外', '平安'],  # 「保幼平安」例外
    '定期險':     ['定期壽險', '定期保險', '一年定期'],
    '微型':       ['微型', '小額終身'],
}


def detect_skip(full_name):
    """從商品全名判斷不轉類別"""
    if '保幼平安' in full_name:  # 例外
        return None
    for category, keywords in SKIP_KEYWORDS.items():
        if any(kw in full_name for kw in keywords):
            return f"❌ {category}"
    return None


def detect_mixed_product(wb):
    """v5 新增:偵測『意外+醫療+生存』混合型(AJI 安康如意)"""
    ws = wb['輸入頁'] if '輸入頁' in wb.sheetnames else None
    if not ws: return None
    
    indicators = set()
    for r in range(5, 20):
        try:
            row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
            text = ' '.join(str(v) for v in row if v)
            if '意外身故' in text: indicators.add('意外')
            if '失能保險金' in text: indicators.add('失能')
            if '燒燙傷' in text: indicators.add('燒燙傷')
            if '住院醫療' in text: indicators.add('醫療')
            if '生存保險金' in text: indicators.add('生存')
        except: pass
    
    # 同時含 意外/失能/燒燙傷 之中至少 2 種 → 混合型
    danger = {'意外', '失能', '燒燙傷'}
    if len(indicators & danger) >= 2:
        return f"⚠️ 混合型保障線索:{indicators}"
    return None


def detect_age_based_period(wb):
    """v5 新增:偵測歲滿期型『繳費至 N 歲』(PAJ 優富年年)"""
    ws = wb['輸入頁'] if '輸入頁' in wb.sheetnames else None
    if not ws: return None
    
    for r in range(20, 50):
        try:
            row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
            text = ' '.join(str(v) for v in row if v)
            m = re.search(r'繳費至\s*(\d+)\s*歲', text)
            if m: return int(m.group(1))
            m = re.search(r'(\d+)\s*歲滿期', text)
            if m: return int(m.group(1))
        except: pass
    return None


def get_full_name(wb):
    """從 R2 取商品全名"""
    ws = wb['輸入頁'] if '輸入頁' in wb.sheetnames else wb[wb.sheetnames[0]]
    for r in range(1, 5):
        try:
            row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
            for v in row:
                if v and isinstance(v, str) and len(v) > 5:
                    return v
        except: pass
    return ''


def detect_engine(wb):
    """偵測使用哪個 engine"""
    sheets = wb.sheetnames
    has_div = any('分紅' in s for s in sheets)
    has_total = '總表' in sheets
    has_print_simple = '列印頁-簡' in sheets
    has_print = '列印頁' in sheets
    
    # 富邦變額型特徵
    if any(s in sheets for s in ['ROP', 'tbULMultiple']) or any('正' in s and '計算內容' in s for s in sheets):
        return None, '❌ 變額/萬能(不支援)'
    
    if has_div and has_total:
        return 'prudential_v2', '🌸 三情境分紅型'
    if has_total and has_print_simple:
        return 'twlife_v1', '📈 標準利變型'
    if has_print and not has_total:
        return 'simple_print_v1', '📋 簡單列印頁(養老/還本)'
    
    return None, '❓ 結構特殊'


def step0_inspect(xlsx_path):
    """Step 0 完整偵察"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    
    full_name = get_full_name(wb)
    skip_check = detect_skip(full_name)
    if skip_check:
        return {
            'engine': None, 'skip': True, 'reason': skip_check,
            'full_name': full_name, 'sheets': wb.sheetnames
        }
    
    mixed_check = detect_mixed_product(wb)
    if mixed_check:
        return {
            'engine': None, 'skip': True, 'reason': '❌ 混合型(意外+醫療+生存)',
            'full_name': full_name, 'sheets': wb.sheetnames, 'mixed': mixed_check
        }
    
    age_based_target = detect_age_based_period(wb)
    
    engine, type_desc = detect_engine(wb)
    
    return {
        'engine': engine, 'skip': engine is None, 'type_desc': type_desc,
        'full_name': full_name, 'sheets': wb.sheetnames,
        'age_based_target': age_based_target,
    }


# =============================================================================
# Step 1:基準參數抽取(v5 強化版)
# =============================================================================

def get_real_currency(wb, base_premium=None, base_sa=None):
    """v5 嚴格 currency:鎖定 R44+ 『計價幣別』標籤,排除法規條文"""
    ws = wb['輸入頁']
    for r in range(40, 60):
        try: row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        except: continue
        for v in row:
            if v is None: continue
            s = str(v)
            if '計價幣別' in s:
                if '美元' in s: return 'USD'
                if '臺幣' in s or '台幣' in s: return 'TWD'
                if '澳幣' in s: return 'AUD'
                if '人民幣' in s: return 'CNY'
    
    # fallback 1:看商品全名
    full_name = get_full_name(wb)
    if '美元' in full_name or '外幣' in full_name: return 'USD'
    if '澳幣' in full_name: return 'AUD'
    if '人民幣' in full_name: return 'CNY'
    
    # 🆕 fallback 2:從 base_premium / base_sa 數量級反推
    # USD 商品 base_premium 通常 100-50萬,TWD 通常 5000-1000萬
    # USD 商品 base_sa 通常 1萬-500萬,TWD 通常 30萬-1億
    if base_premium and base_sa:
        # 主要看 base_sa(更穩定)
        if base_sa >= 5_000_000:  # >= 500萬,大概率 TWD
            return 'TWD'
        if base_sa <= 200_000 and base_premium <= 100_000:  # USD 可能
            # 但 50000 USD 跟 50000 TWD 都可能 → 看 base_premium / base_sa 比
            # USD 商品 prem/sa 約 0.05-0.5;TWD 商品 prem/sa 約 0.005-0.1
            ratio = base_premium / base_sa
            if 0.03 < ratio < 0.6:
                return 'USD'
    
    return 'TWD'  # 富邦預設


def extract_base_sa_strict(ws):
    """v5 嚴格 base_sa:鎖定『主約險種代號』下方『保額(萬)』"""
    for r in range(30, 50):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        for i, v in enumerate(row):
            if v is None: continue
            if '主約險種代號' in str(v):
                # 在當列或下兩列找『保額(萬)』
                for r2 in range(r, min(r+3, ws.max_row + 1)):
                    row2 = list(ws.iter_rows(min_row=r2, max_row=r2, values_only=True))[0]
                    for i2, v2 in enumerate(row2):
                        if v2 is None: continue
                        s2 = str(v2)
                        if '保額(萬)' in s2 or '保額\uff08萬\uff09' in s2:
                            for j in range(i2+1, min(i2+6, len(row2))):
                                val = row2[j]
                                if isinstance(val, (int, float)) and 0 < val <= 100000:
                                    return int(val * 10000), 10000
                return None, None
    
    # fallback:全表掃描(較不嚴格)
    for r in range(1, 50):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        for i, v in enumerate(row):
            if v is None: continue
            s = str(v)
            if '基本保額' in s:
                val, _ = find_value_after(row, i, value_filter=lambda x: isinstance(x, (int, float)) and x > 100)
                if val: return int(val), 1
    return None, None


def extract_base_premium(ws):
    """v5 區分繳費型態抽 base_premium"""
    premium_net, premium_gross = None, None
    
    for r in range(1, 60):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        for i, v in enumerate(row):
            if v is None: continue
            s = str(v)
            
            if '折扣後' in s and '保費' in s:
                val, _ = find_value_after(row, i,
                    value_filter=lambda x: isinstance(x, (int, float)) and x > 100)
                if val and (premium_net is None or val < premium_net):
                    premium_net = int(val)
            
            if ('年繳保費' in s or '躉繳保費' in s) and '折扣' not in s:
                val, _ = find_value_after(row, i,
                    value_filter=lambda x: isinstance(x, (int, float)) and x > 100)
                if val and premium_gross is None:
                    premium_gross = int(val)
    
    # 躉繳沒折扣後 fallback
    if premium_net is None and premium_gross is not None:
        premium_net = premium_gross
    
    return premium_net, premium_gross


def extract_period(ws, base_age):
    """v5 period:標準 + 歲滿期(支援字串數字 / 「歲滿期」標籤)"""
    for r in range(1, 60):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        for i, v in enumerate(row):
            if v is None: continue
            s = str(v)
            
            if ('主約繳費年期' in s or '繳費年期' in s):
                # 🆕 標準型:int 或 字串
                for j in range(i+1, min(i+6, len(row))):
                    val = row[j]
                    if val is None: continue
                    
                    # int 直接抓
                    if isinstance(val, int) and 1 <= val <= 30:
                        # 但要確認後面有沒有「歲滿期」字樣
                        rest_text = ' '.join(str(x) for x in row[j+1:j+5] if x is not None)
                        if '歲滿期' in rest_text and base_age is not None:
                            return val - base_age, 'age_based', val
                        return int(val), 'fixed', None
                    
                    # 🆕 字串數字 + 後面有「歲滿期」
                    if isinstance(val, str):
                        try:
                            num = int(val)
                        except:
                            continue
                        if 1 <= num <= 30:
                            return num, 'fixed', None
                        # 🆕 大數字(如 55) + 後面有「歲滿期」
                        rest_text = ' '.join(str(x) for x in row[j+1:j+5] if x is not None)
                        if '歲滿期' in rest_text and base_age is not None and 10 <= num <= 100:
                            return num - base_age, 'age_based', num
            
            # 「繳費至 N 歲」
            m = re.search(r'繳費至\s*(\d+)\s*歲', s)
            if m and base_age is not None:
                target_age = int(m.group(1))
                return target_age - base_age, 'age_based', target_age
            
            # 「N 歲滿期」
            m = re.search(r'(\d+)\s*歲滿期', s)
            if m and base_age is not None:
                target_age = int(m.group(1))
                return target_age - base_age, 'age_based', target_age
    
    return None, None, None


def extract_input_page(wb):
    """v5 完整輸入頁抽取"""
    ws = wb['輸入頁']
    
    info = {
        'sex': None, 'age': None, 'sa': None, 'sa_unit': 1, 
        'period': None, 'period_type': None, 'period_target_age': None,
        'premium_net': None, 'premium_gross': None,
        'declared_rate': None, 'currency': None, 'dividend_option': None,
        'min_sa': None, 'max_sa': None, 'max_age': None,
    }
    
    # 第一輪:抓 sex/age
    for r in range(1, min(80, ws.max_row + 1)):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        for i, v in enumerate(row):
            if v is None: continue
            s = str(v)
            
            if info['sex'] is None and v in ('男', '女'):
                for j in range(max(0, i-3), i):
                    if row[j] and '性別' in str(row[j]):
                        info['sex'] = 'M' if v == '男' else 'F'
                        break
            
            if info['age'] is None and '保險年齡' in s:
                val, _ = find_value_after(row, i,
                    value_filter=lambda x: isinstance(x, int) and 0 <= x <= 110)
                if val is not None: info['age'] = int(val)
    
    # v5 嚴格 base_sa
    sa_result = extract_base_sa_strict(ws)
    if sa_result[0]:
        info['sa'] = sa_result[0]
        info['sa_unit'] = sa_result[1]
    
    # v5 base_premium
    info['premium_net'], info['premium_gross'] = extract_base_premium(ws)
    
    # v5 period
    info['period'], info['period_type'], info['period_target_age'] = extract_period(ws, info['age'])
    
    # 🆕 currency 在 base_sa/premium 抽完後判斷(才能用數量級反推)
    info['currency'] = get_real_currency(wb, base_premium=info['premium_net'], base_sa=info['sa'])
    
    # 其他欄位
    for r in range(1, min(80, ws.max_row + 1)):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        for i, v in enumerate(row):
            if v is None: continue
            s = str(v)
            
            if info['declared_rate'] is None and '宣告利率' in s:
                m = re.search(r'(\d+\.?\d*)\s*%', s)
                if m: info['declared_rate'] = float(m.group(1)) / 100
                else:
                    val, _ = find_value_after(row, i,
                        value_filter=lambda x: isinstance(x, (int, float)) and 0 < x < 0.2)
                    if val is not None: info['declared_rate'] = float(val)
            
            if '紅利給付方式' in s or '增值回饋' in s:
                val, _ = find_value_after(row, i)
                if val:
                    s2 = str(val)
                    if '儲存' in s2: info['dividend_option'] = '儲存生息'
                    elif '繳清' in s2 or '增繳' in s2: info['dividend_option'] = '繳清保險金額'
                    elif '現金' in s2: info['dividend_option'] = '現金給付'
            
            if '保額限制' in s or ('保額' in s and '~' in s):
                m = re.search(r'([\d,.]+)\s*~\s*([\d,.]+)', s)
                if m and info['sa_unit']:
                    info['min_sa'] = int(float(m.group(1).replace(',', '')) * info['sa_unit'])
                    info['max_sa'] = int(float(m.group(2).replace(',', '')) * info['sa_unit'])
            
            if '投保年齡' in s:
                m = re.search(r'(\d+)歲\s*~\s*(\d+)歲', s)
                if m: info['max_age'] = int(m.group(2))
    
    return info


# =============================================================================
# Step 2:逐年表抽取(三引擎)
# =============================================================================

def find_column_indices(ws, header_search_rows=None):
    """v5 動態欄位偵測"""
    if header_search_rows is None:
        header_search_rows = range(1, 22)
    
    titles = {}
    for r in header_search_rows:
        try: row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        except: continue
        for i, v in enumerate(row):
            if v is None: continue
            s = str(v).replace('\n', ' ')
            if i not in titles: titles[i] = ''
            titles[i] += ' ' + s
    
    cols = {}
    for i, t in titles.items():
        if 'cum_prem' not in cols and ('累計實繳' in t or '累計所繳' in t):
            cols['cum_prem'] = i
        if 'death_benefit' not in cols and '身故' in t:
            if any(kw in t for kw in ['可領總金額', '+ C', '+C', '含']):
                cols['death_benefit'] = i
            elif '保障' in t and '可領' not in t:
                cols['death_benefit'] = i
        if 'cv_total' not in cols:
            if '解約' in t and any(kw in t for kw in ['可領總金額', '+ C', '+C', '含']):
                cols['cv_total'] = i
            elif '解約金' in t and '減額' not in t:
                cols['cv_total'] = i
        if 'dividend_year' not in cols and '增值回饋分享金' in t and '累計' not in t:
            cols['dividend_year'] = i
        if 'dividend_cum' not in cols and '累計' in t and '增值回饋' in t:
            cols['dividend_cum'] = i
        if 'increment_amount' not in cols and '累計增額繳清' in t:
            cols['increment_amount'] = i
        if 'survival_year' not in cols and '生存保險金' in t and '累計' not in t:
            cols['survival_year'] = i
        if 'survival_cum' not in cols and '累計' in t and '生存保險金' in t:
            cols['survival_cum'] = i
    return cols


def extract_twlife_v1(wb, info):
    """twlife_v1:標準利變型(總表 + 列印頁-簡)"""
    if '總表' not in wb.sheetnames or '列印頁-簡' not in wb.sheetnames:
        return None, "缺總表 或 列印頁-簡"
    
    ws_total = wb['總表']
    ws_print = wb['列印頁-簡']
    
    cols = find_column_indices(ws_print)
    if 'cv_total' not in cols or 'death_benefit' not in cols:
        return None, f"列印頁-簡找不到欄位 ({cols})"
    
    total_rows = extract_continuous_y(ws_total, 4)
    print_rows = extract_continuous_y(ws_print, 18)
    if not total_rows or not print_rows:
        return None, "資料抽不到"
    
    sample = total_rows[0][1]
    has_gross_col = len(sample) > 27 and isinstance(safe_get(sample, 27), (int, float))
    
    total_data = {}
    for r, row in total_rows:
        cum_4 = safe_get(row, 4)
        cum_27 = safe_get(row, 27) if has_gross_col else None
        total_data[row[1]] = {
            "age": int(row[2]),
            "cv_basic": int(round(safe_get(row, 8) or 0)),
            "cum_gross": int(round(cum_27 or cum_4 or 0)),
            "cum_net": int(round(cum_4 or 0)),
        }
    
    print_data = {}
    for r, row in print_rows:
        rec = {
            "death_benefit": int(round(safe_get(row, cols['death_benefit']) or 0)),
            "cv_total": int(round(safe_get(row, cols['cv_total']) or 0)),
        }
        if 'dividend_year' in cols:
            rec['dividend_year'] = int(round(safe_get(row, cols['dividend_year']) or 0))
        if 'dividend_cum' in cols:
            rec['dividend_cum'] = int(round(safe_get(row, cols['dividend_cum']) or 0))
        if 'increment_amount' in cols:
            rec['increment_amount'] = int(round(safe_get(row, cols['increment_amount']) or 0))
        print_data[row[1]] = rec
    
    schedule = []
    for y in sorted(total_data.keys()):
        if y not in print_data: continue
        t = total_data[y]; p = print_data[y]
        rec = {
            "y": int(y), "age": t["age"],
            "cum_prem": t["cum_gross"],
            "cv_basic": t["cv_basic"],
            "cv_total": p["cv_total"],
            "death_benefit": p["death_benefit"],
        }
        if t["cum_net"] != t["cum_gross"]: rec["cum_prem_net"] = t["cum_net"]
        if p.get("dividend_year"): rec["dividend_year"] = p["dividend_year"]
        if p.get("dividend_cum"): rec["dividend_cum"] = p["dividend_cum"]
        if p.get("increment_amount"): rec["increment_amount"] = p["increment_amount"]
        schedule.append(rec)
    schedule.sort(key=lambda r: r['y'])
    
    return schedule, None


def extract_prudential_v2(wb, info):
    """prudential_v2:三情境分紅(總表 + 總表_分紅_M/L)"""
    if '總表' not in wb.sheetnames:
        return None, "缺總表"
    if '總表_分紅_M' not in wb.sheetnames or '總表_分紅_L' not in wb.sheetnames:
        return None, "缺總表_分紅_M/L"
    
    ws_total = wb['總表']
    none_rows = extract_continuous_y(ws_total, 4)
    if not none_rows: return None, "總表抽不到"
    
    mid_rows = extract_continuous_y(wb['總表_分紅_M'], 5)
    low_rows = extract_continuous_y(wb['總表_分紅_L'], 5)
    
    schedule = []
    for r, row in none_rows:
        cum = safe_get(row, 4); A = safe_get(row, 7); B = safe_get(row, 8)
        if cum is None or A is None or B is None: continue
        schedule.append({
            "y": int(row[1]), "age": int(row[2]),
            "cum_prem": int(round(cum)),
            "cv_basic": int(round(B)), "cv_total": int(round(B)),
            "death_benefit": int(round(A)),
        })
    
    def rows_to_dict(rows):
        d = {}
        for r, row in rows:
            d[row[1]] = {
                "dividend_year": int(round(safe_get(row, 10) or 0)),
                "db_with_dividend": int(round(safe_get(row, 22) or 0)),
                "cv_total": int(round(safe_get(row, 23) or 0)),
            }
        return d
    
    mid_data = rows_to_dict(mid_rows)
    low_data = rows_to_dict(low_rows)
    
    for rec in schedule:
        y = rec['y']
        rec['scenarios'] = {
            "none": {"dividend_year": 0, "db_with_dividend": rec['death_benefit'], "cv_total": rec['cv_total']},
            "mid":  mid_data.get(y, {"dividend_year": 0, "db_with_dividend": rec['death_benefit'], "cv_total": rec['cv_total']}),
            "low":  low_data.get(y, {"dividend_year": 0, "db_with_dividend": rec['death_benefit'], "cv_total": rec['cv_total']}),
        }
    schedule.sort(key=lambda r: r['y'])
    
    return schedule, None


def extract_simple_print_v1(wb, info):
    """simple_print_v1:養老/還本特殊(只有列印頁)"""
    if '列印頁' not in wb.sheetnames:
        return None, "缺列印頁"
    
    ws = wb['列印頁']
    
    # 🆕 v5 簡單列印頁的標題列範圍要更寬,因為 R13-17 可能是介紹文,真表頭在 R18-22
    cols = find_column_indices(ws, header_search_rows=range(13, 24))
    
    if 'cum_prem' not in cols or 'death_benefit' not in cols:
        return None, f"列印頁找不到欄位 ({cols})"
    
    # 🆕 找資料起始列(從表頭往下找第一個 y=1 的列)
    data_start = 15
    for r in range(15, min(30, ws.max_row + 1)):
        try:
            row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
            if isinstance(safe_get(row, 1), int) and safe_get(row, 1) == 1:
                data_start = r
                break
        except: pass
    
    schedule = []
    prev_y = 0
    for r in range(data_start, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        y = safe_get(row, 1)
        if not (isinstance(y, int) and 1 <= y <= 110): continue
        age = safe_get(row, 2)
        if not (isinstance(age, int) and 0 <= age <= 110): continue
        if y < prev_y: break
        
        cum = safe_get(row, cols['cum_prem'])
        db = safe_get(row, cols['death_benefit'])
        cv = safe_get(row, cols.get('cv_total', -1)) or 0
        
        if cum is None or db is None: continue
        if cv is None: cv = 0
        
        rec = {"y": int(y), "age": int(age),
               "cum_prem": int(round(cum)),
               "cv_basic": int(round(cv)), "cv_total": int(round(cv)),
               "death_benefit": int(round(db))}
        
        if 'survival_year' in cols:
            sv = safe_get(row, cols['survival_year'])
            if sv: rec['survival_year'] = int(round(sv))
        if 'survival_cum' in cols:
            sc = safe_get(row, cols['survival_cum'])
            if sc: rec['survival_cum'] = int(round(sc))
        if 'increment_amount' in cols:
            ia = safe_get(row, cols['increment_amount'])
            if ia: rec['increment_amount'] = int(round(ia))
        
        schedule.append(rec)
        prev_y = y
    schedule.sort(key=lambda r: r['y'])
    
    return schedule, None


# =============================================================================
# Step 4:Verify(v5 分級警告)
# =============================================================================

def verify(data):
    """v5 自洽性驗證 + warning 分級"""
    sched = data['schedule']
    base = data['meta']
    errors = []
    warnings = {'red': [], 'yellow': [], 'green': []}
    p = base.get('period') or 6
    
    # 缺值
    for k in ['base_sa', 'base_premium', 'period']:
        if not base.get(k): errors.append(f"缺 {k}")
    if base.get('base_age') is None: errors.append("缺 base_age")
    if not base.get('base_sex'): errors.append("缺 base_sex")
    if errors: return errors, warnings
    if not sched: return ["schedule 空"], warnings
    
    # 1. Y1 cum
    if abs(sched[0]['cum_prem'] - base['base_premium']) > 1:
        errors.append(f"1.Y1 cum={sched[0]['cum_prem']} ≠ {base['base_premium']}")
    
    # 1b. 🆕 Y1 vs Y2 差異
    if len(sched) >= 2 and base['base_premium']:
        y1_y2_diff = sched[1]['cum_prem'] - sched[0]['cum_prem']
        expected = base['base_premium']
        if expected > 0 and abs(y1_y2_diff - expected) / expected > 0.01:
            warnings['red'].append(
                f"1b.Y1 vs Y2 cum 差 {y1_y2_diff} ≠ {expected}(可能附約保費分離,建議 SKIP)"
            )
    
    # 2. Y(period) cum
    if len(sched) >= p:
        expected = base['base_premium'] * p
        if abs(sched[p-1]['cum_prem'] - expected) > p:
            errors.append(f"2.Y{p} cum={sched[p-1]['cum_prem']} ≠ {expected}")
    
    # 3. Y(period+1)+ cum 停止
    if len(sched) > p:
        if abs(sched[p]['cum_prem'] - sched[p-1]['cum_prem']) > 1:
            errors.append(f"3.Y{p+1} cum 應停止")
    
    # 4. cv_total >= cv_basic
    fail = next((r for r in sched if r['cv_total'] < r['cv_basic'] - 1), None)
    if fail: errors.append(f"4.Y{fail['y']} cv_total<cv_basic")
    
    # 5. 繳費期內 cv 遞增
    fail = None
    for i in range(1, min(p, len(sched))):
        if base['engine'] == 'prudential_v2':
            curr = sched[i]['scenarios']['mid']['cv_total']
            prev = sched[i-1]['scenarios']['mid']['cv_total']
        else:
            curr = sched[i]['cv_total']; prev = sched[i-1]['cv_total']
        if curr < prev - 1: fail = sched[i]; break
    if fail:
        warnings['yellow'].append(f"5.繳費期 Y{fail['y']} cv 下降")
    
    # 6. Y1 db ≈ base_sa(分級)
    ratio = sched[0]['death_benefit'] / base['base_sa']
    if not 0.95 <= ratio <= 1.05:
        if base.get('base_age', -1) <= 5:
            warnings['green'].append(f"6.Y1 db/sa={ratio:.3f}(0-5 歲法規 ramp 正常)")
        elif base['base_premium'] / base['base_sa'] > 0.5:
            warnings['green'].append(f"6.Y1 db/sa={ratio:.3f}(保費/保額比高,正常)")
        else:
            warnings['red'].append(f"6.Y1 db/sa={ratio:.3f}(可能 ramp_up,要查 DM)")
    
    # 7. 末年齡
    last_age = sched[-1].get('age')
    if last_age and last_age > 110: errors.append(f"7.末年齡={last_age}")
    
    # 8. 筆數
    if len(sched) < 30:
        warnings['yellow'].append(f"8.筆數 {len(sched)} 過少(可能養老/滿期型)")
    elif len(sched) < 50:
        warnings['green'].append(f"8.筆數 {len(sched)}")
    
    # 9. discount(discount=0 跳過)
    if 'base_premium_gross' in base and base.get('discount', 0) > 0:
        gross = base['base_premium_gross']; net = base['base_premium']
        d = base['discount']
        expected_net = gross * (1 - d)
        if abs(expected_net - net) > 1:
            errors.append(f"9.discount 不自洽")
    
    # prudential_v2 額外
    if base['engine'] == 'prudential_v2':
        if len(sched) > p:
            mid_y_after = sched[p].get('scenarios', {}).get('mid', {}).get('dividend_year', 0)
            if mid_y_after == 0:
                warnings['red'].append(f"11.Y{p+1} mid div_y=0(可能抽錯欄位)")
        
        fail = None
        for r in sched:
            mid_db = r['scenarios']['mid']['db_with_dividend']
            if mid_db < r['death_benefit'] - 1: fail = r; break
        if fail: errors.append(f"12.Y{fail['y']} mid.db<death_benefit")
    
    # 14. 🆕 currency × base_sa 交叉驗證
    sa = base['base_sa']
    cur = base['currency']
    if cur == 'USD' and sa > 100_000_000:
        warnings['red'].append(f"14.USD sa={sa:,} 異常大,可能 currency 抽錯")
    if cur == 'TWD' and sa < 50_000:
        warnings['red'].append(f"14.TWD sa={sa:,} 異常小,可能 currency 抽錯")
    
    return errors, warnings


# =============================================================================
# 主入口
# =============================================================================

MID_DIVIDEND_DEFAULT = {
    'USD': 0.0550, 'TWD': 0.0450,
    'AUD': 0.0500, 'CNY': 0.0400,
}


def process_one_file(xlsx_path, plan_code, company, product_name, engine='auto'):
    """
    處理單一檔案
    
    Returns:
        (data, errors, warnings) — data 是 None 表示失敗
    """
    xlsx_path = Path(xlsx_path)
    
    # Step 0
    inspect = step0_inspect(xlsx_path)
    if inspect.get('skip'):
        return None, [inspect.get('reason', 'SKIP')], {'red': [], 'yellow': [], 'green': []}
    
    # 自動偵測 engine
    if engine == 'auto':
        engine = inspect['engine']
    
    if engine is None:
        return None, ['無法偵測 engine'], {'red': [], 'yellow': [], 'green': []}
    
    # Step 1
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    info = extract_input_page(wb)
    
    # Step 2
    if engine == 'twlife_v1':
        schedule, err = extract_twlife_v1(wb, info)
    elif engine == 'prudential_v2':
        schedule, err = extract_prudential_v2(wb, info)
    elif engine == 'simple_print_v1':
        schedule, err = extract_simple_print_v1(wb, info)
        engine = 'twlife_v1'  # simple_print 用 twlife_v1 schema
    else:
        return None, [f'未知 engine: {engine}'], {'red': [], 'yellow': [], 'green': []}
    
    if err:
        return None, [err], {'red': [], 'yellow': [], 'green': []}
    
    # 組 meta
    discount = 0
    if info['premium_gross'] and info['premium_net'] and info['premium_gross'] != info['premium_net']:
        discount = round(1 - info['premium_net'] / info['premium_gross'], 4)
    
    cur = info['currency'] or 'USD'
    meta = {
        "product_id": plan_code,
        "company": company,
        "product_name": product_name,
        "currency": cur,
        "period": info['period'],
        "engine": engine,
        "base_sex": info['sex'],
        "base_age": info['age'],
        "base_sa": info['sa'],
        "base_premium": info['premium_net'],
        "discount": discount,
        "declared_rate": info['declared_rate'] or 0 if engine != 'prudential_v2' else 0,
        "source_file": xlsx_path.name,
        "extracted_at": date.today().isoformat(),
    }
    if info['premium_gross']:
        meta['base_premium_gross'] = info['premium_gross']
    if info['dividend_option']:
        meta['dividend_option'] = info['dividend_option']
    if engine == 'prudential_v2':
        meta['mid_dividend_rate'] = MID_DIVIDEND_DEFAULT.get(cur, 0.055)
    if info['period_type'] == 'age_based':
        meta['period_type'] = 'age_based'
        meta['period_target_age'] = info['period_target_age']
    
    data = {"meta": meta, "schedule": schedule}
    
    # Step 4
    errors, warnings = verify(data)
    
    return data, errors, warnings


def process_batch(tasks, output_dir='/mnt/user-data/outputs'):
    """
    批次處理
    
    tasks: list of (plan_code, company, product_name, xlsx_path, engine)
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)
    
    results = []
    for code, company, name, xlsx_path, engine in tasks:
        try:
            data, errors, warnings = process_one_file(xlsx_path, code, company, name, engine)
            
            if data is None:
                results.append({
                    'code': code, 'name': name, 'status': 'FAIL',
                    'errors': errors, 'warnings': warnings,
                })
                continue
            
            if errors:
                results.append({
                    'code': code, 'name': name, 'status': 'VERIFY_FAIL',
                    'errors': errors, 'warnings': warnings,
                    'meta': data['meta'],
                })
                continue
            
            # 寫檔
            out = output_dir / f'{code}.json'
            with open(out, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            results.append({
                'code': code, 'name': name, 'status': 'OK',
                'errors': [], 'warnings': warnings,
                'meta': data['meta'],
                'sched_count': len(data['schedule']),
                'out_path': str(out),
            })
        except Exception as e:
            import traceback
            results.append({
                'code': code, 'name': name, 'status': 'EXCEPTION',
                'msg': str(e), 'tb': traceback.format_exc()[:500],
            })
    
    return results


# =============================================================================
# 範例:單檔測試
# =============================================================================
if __name__ == '__main__':
    import sys
    
    if len(sys.argv) < 5:
        print("用法: python excel_to_json_v5.py <xlsx_path> <plan_code> <company> <product_name>")
        print("範例: python excel_to_json_v5.py /path/PFA.xlsx PFA 富邦人壽 美富紅運外幣分紅終身壽險")
        sys.exit(1)
    
    xlsx, code, company, name = sys.argv[1:5]
    
    print(f"=== 處理 {code} ({name}) ===")
    data, errors, warnings = process_one_file(xlsx, code, company, name)
    
    if data is None:
        print(f"❌ 失敗: {errors}")
        sys.exit(1)
    
    print(f"✅ 成功!")
    print(f"  engine: {data['meta']['engine']}")
    print(f"  currency: {data['meta']['currency']}")
    print(f"  base_sa: {data['meta']['base_sa']:,}")
    print(f"  base_premium: {data['meta']['base_premium']:,}")
    print(f"  period: {data['meta']['period']}")
    print(f"  schedule: {len(data['schedule'])} 筆")
    
    if warnings['red']:
        print(f"\n⚠️🔴 嚴重警告 ({len(warnings['red'])}):")
        for w in warnings['red']: print(f"  - {w}")
    if warnings['yellow']:
        print(f"\n⚠️🟡 普通警告 ({len(warnings['yellow'])}):")
        for w in warnings['yellow']: print(f"  - {w}")
    if errors:
        print(f"\n❌ 錯誤:")
        for e in errors: print(f"  - {e}")
